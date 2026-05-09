"""
Voice of Field — Weekly Trend Analysis
=======================================
Reads the "Voice of Field (Weekly Insights).xlsx" workbook and produces
trend data across weeks for four dimensions:
  - Category
  - Product Area
  - Product Subcategory
  - Sales Stage

Trend methods produced:
  1. Raw weekly counts
  2. 4-week moving average
  3. Week-over-week % change
  4. Share-of-voice (% of week total)
  5. Momentum score (recent 4-wk avg ÷ prior 4-wk avg)
  6. New / disappeared accounts per week

Output: an Excel workbook with one tab per analysis type,
        plus a console summary of the most notable movements.
"""

from __future__ import annotations
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_PATH = Path("data/inputs/Voice of Field (Weekly Insights).xlsx")
OUTPUT_PATH = Path("data/outputs/vof_weekly_trends.xlsx")

DIMENSIONS = ["Category", "Product Area", "Product Subcategory", "Sales Stage"]

MOVING_AVG_WINDOW = 4

# Canonical field names mapped from the various header spellings found across sheets.
FIELD_ALIASES = {
    "account":              ["account", "account name"],
    "insight":              ["insight", "insights"],
    "category":             ["category", "categories", "insight category"],
    "area":                 ["product area", "product areas"],
    "subcat":               ["product subcategory", "subcategory"],
    "stage":                ["sales stage", "sales stage (now)"],
    "opp_amount":           ["opportunity $ amount", "opportunity amount (usd)",
                             "opportunity amount ($)", "opportunity $ amount"],
}


# ---------------------------------------------------------------------------
# 1. Parse sheet names into sortable dates
# ---------------------------------------------------------------------------

def parse_sheet_date(name: str) -> datetime:
    """Sheet names like '3926' -> March 9, 2026; '111725' -> Nov 17, 2025."""
    name = name.strip()
    if len(name) <= 4:
        month = name[0]
        day = name[1:3]
        year = "20" + name[3:]
    elif len(name) == 5:
        if int(name[:2]) > 12:
            month = name[0]
            day = name[1:3]
            year = "20" + name[3:]
        else:
            month = name[:2]
            day = name[2:3]
            year = "20" + name[3:]
    else:
        month = name[:2]
        day = name[2:4]
        year = "20" + name[4:]
    return datetime(int(year), int(month), int(day))


def parse_sheet_date_robust(name: str) -> datetime:
    """
    Sheet names are M/D/YY with slashes removed.
    Year is always the last 2 chars. Remaining prefix is M + D where
    month is 1–12 and day is 1–31, with no zero-padding guarantee.
    We try M=1-digit then M=2-digit, picking the first valid date.
    """
    n = name.strip()
    year = 2000 + int(n[-2:])
    prefix = n[:-2]  # month + day concatenated

    candidates: list[tuple[int, int]] = []
    if len(prefix) >= 2:
        candidates.append((int(prefix[0]), int(prefix[1:])))
    if len(prefix) >= 3:
        candidates.append((int(prefix[:2]), int(prefix[2:])))

    for m, d in candidates:
        try:
            dt = datetime(year, m, d)
            return dt
        except ValueError:
            continue

    raise ValueError(f"Cannot parse sheet name '{name}' as a date")


# ---------------------------------------------------------------------------
# 2. Load data
# ---------------------------------------------------------------------------

def _normalize_dimension_value(dim_name: str, val: str) -> str:
    """Collapse spelling/casing/punctuation variants into canonical buckets."""
    val = val.strip().rstrip(".")
    lower = val.lower()

    if dim_name == "Category":
        if "blocker" in lower:
            return "Customer Requirements (Blocker)"
        if "enhancement" in lower or lower == "customer requirements":
            return "Customer Requirements (Enhancement)"
        if "customer requirements" in lower and "blocker" not in lower:
            return "Customer Requirements (Enhancement)"
        if lower.startswith("capacity"):
            return "Capacity"
        if "education" in lower:
            return "Education Gaps"
        if "issue" in lower:
            return "Issues"
        if "pricing" in lower or "terms" in lower:
            return "Pricing/Terms"
        if lower in ("null", "other", "multiple", "partnership", "performance"):
            return lower.title()
        return val

    if dim_name == "Sales Stage":
        if "closed won" in lower:
            return "Closed Won"
        if "closed lost" in lower:
            return "Closed Lost"
        if "discovery" in lower:
            return "Discovery"
        if "technical evaluation" in lower or "tech eval" in lower:
            return "Technical Evaluation"
        if "capacity review" in lower or "poc" in lower:
            return "Capacity Review"
        if "negotiation" in lower:
            return "Negotiations"
        if "legal" in lower:
            return "Legal Redlines"
        if "proposal" in lower or "price quote" in lower:
            return "Proposal/Price Quote"
        if "qualification" in lower or "evaluation" in lower:
            return "Qualification"
        if "pipeline" in lower:
            return "Pipeline"
        if "prospect" in lower:
            return "Prospect"
        if "current customer" in lower:
            return "Current Customer"
        if any(x in lower for x in [
            "n/a", "no open opp", "no active opp", "not found",
            "no open opportunity", "former customer", "pre\u2011opportunity",
            "pre-opportunity", "account present", "unknown", "na",
            "opp stage not re", "stage from main", "no opp",
        ]):
            return "n/a"
        return val

    if dim_name == "Product Subcategory":
        if any(x in lower for x in ["sunk", "training"]):
            return "Training"
        if "console" in lower or "api" in lower or "terraform" in lower:
            return "Console / API / Terraform"
        if "consumption" in lower:
            return "Consumption Models"
        if lower in ("sunk/ training", "sunk/training", "sunk"):
            return "Training"
        return val

    if dim_name == "Product Area":
        if lower in ("ai services (sunk/inference)", "ai services"):
            return "AI Services"
        if lower in ("weights & biases", "w&b"):
            return "Weights & Biases"
        if lower in ("other (cost)", "other (term)"):
            return "Other"
        return val

    return val


def _resolve_header_map(header_cells: list[str | None]) -> dict[str, int]:
    """
    Given a row of header cell values, return {canonical_field: col_index}
    by matching against FIELD_ALIASES (case-insensitive).
    """
    mapping: dict[str, int] = {}
    for idx, raw_val in enumerate(header_cells):
        if raw_val is None:
            continue
        normed = raw_val.strip().lower()
        for field, aliases in FIELD_ALIASES.items():
            if normed in aliases:
                mapping[field] = idx
                break
    return mapping


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    """Scan the first 5 rows for one that looks like a header (has 'account')."""
    for row_idx in range(1, 6):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        cell_strs = [str(c).strip().lower() if c else "" for c in cells]
        if any(a in cell_strs for a in FIELD_ALIASES["account"]):
            header_map = _resolve_header_map(
                [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            )
            return row_idx, header_map
    raise ValueError(f"Could not find header row in sheet '{ws.title}'")


def load_weekly_data(path: Path) -> dict:
    """
    Returns:
        weeks: list of (date, label) sorted chronologically
        dimension_counts: {dim_name: {week_label: {value: count}}}
        account_sets: {week_label: set(account_names)}
        week_totals: {week_label: int}
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    raw: list[tuple[datetime, str, list[dict]]] = []
    for sheet_name in wb.sheetnames:
        dt = parse_sheet_date_robust(sheet_name)
        ws = wb[sheet_name]

        header_row, col_map = _find_header_row(ws)
        if "account" not in col_map:
            continue

        rows = []
        for row in ws.iter_rows(min_row=header_row + 1, max_col=ws.max_column, values_only=True):
            acct_idx = col_map["account"]
            if acct_idx >= len(row) or row[acct_idx] is None:
                break
            record = {}
            for field in ["account", "insight", "category", "area", "subcat", "stage", "opp_amount"]:
                idx = col_map.get(field)
                if idx is not None and idx < len(row):
                    record[field] = str(row[idx]).strip() if row[idx] is not None else ""
                else:
                    record[field] = ""
            rows.append(record)
        raw.append((dt, sheet_name, rows))

    raw.sort(key=lambda x: x[0])

    weeks = [(dt, dt.strftime("%b %d, %Y")) for dt, _, _ in raw]
    week_labels = [w[1] for w in weeks]

    dim_field_map = {
        "Category": "category",
        "Product Area": "area",
        "Product Subcategory": "subcat",
        "Sales Stage": "stage",
    }

    dim_counts: dict[str, dict[str, dict[str, int]]] = {
        d: defaultdict(lambda: defaultdict(int)) for d in DIMENSIONS
    }
    account_sets: dict[str, set[str]] = {}
    week_totals: dict[str, int] = {}

    for dt, sheet_name, rows in raw:
        label = dt.strftime("%b %d, %Y")
        accounts = set()
        week_totals[label] = len(rows)
        for r in rows:
            accounts.add(r["account"])
            for dim_name, field in dim_field_map.items():
                val = r.get(field, "").strip()
                val = _normalize_dimension_value(dim_name, val) if val else "(blank)"
                dim_counts[dim_name][label][val] += 1
        account_sets[label] = accounts

    return {
        "weeks": weeks,
        "week_labels": week_labels,
        "dim_counts": dim_counts,
        "account_sets": account_sets,
        "week_totals": week_totals,
    }


# ---------------------------------------------------------------------------
# 3. Compute trend metrics
# ---------------------------------------------------------------------------

def compute_raw_counts(dim_counts, week_labels, dim_name):
    """Returns {value: [count_per_week]}."""
    all_values = sorted({
        v for wk in week_labels for v in dim_counts[dim_name][wk]
    })
    result = {}
    for v in all_values:
        result[v] = [dim_counts[dim_name][wk].get(v, 0) for wk in week_labels]
    return result


def compute_moving_average(raw_counts, window=MOVING_AVG_WINDOW):
    """Returns {value: [ma_per_week]} with None for early weeks."""
    result = {}
    for v, counts in raw_counts.items():
        ma = []
        for i in range(len(counts)):
            if i < window - 1:
                ma.append(None)
            else:
                ma.append(sum(counts[i - window + 1 : i + 1]) / window)
        result[v] = ma
    return result


def compute_wow_change(raw_counts):
    """Week-over-week % change. None for first week."""
    result = {}
    for v, counts in raw_counts.items():
        changes = [None]
        for i in range(1, len(counts)):
            prev = counts[i - 1]
            curr = counts[i]
            if prev == 0:
                changes.append(None if curr == 0 else float("inf"))
            else:
                changes.append(round((curr - prev) / prev * 100, 1))
        result[v] = changes
    return result


def compute_share_of_voice(dim_counts, week_labels, dim_name, week_totals):
    """Each value's count as % of total insights that week."""
    all_values = sorted({
        v for wk in week_labels for v in dim_counts[dim_name][wk]
    })
    result = {}
    for v in all_values:
        shares = []
        for wk in week_labels:
            total = week_totals[wk]
            cnt = dim_counts[dim_name][wk].get(v, 0)
            shares.append(round(cnt / total * 100, 1) if total else 0)
        result[v] = shares
    return result


def compute_momentum(raw_counts, window=MOVING_AVG_WINDOW):
    """
    Ratio of recent N-week avg to prior N-week avg.
    > 1.0 = accelerating, < 1.0 = decelerating.
    Only available when 2*window weeks exist.
    """
    result = {}
    for v, counts in raw_counts.items():
        scores = []
        for i in range(len(counts)):
            if i < 2 * window - 1:
                scores.append(None)
            else:
                recent = sum(counts[i - window + 1 : i + 1]) / window
                prior = sum(counts[i - 2 * window + 1 : i - window + 1]) / window
                if prior == 0:
                    scores.append(None if recent == 0 else float("inf"))
                else:
                    scores.append(round(recent / prior, 2))
        result[v] = scores
    return result


def compute_account_churn(account_sets, week_labels):
    """Returns (new_accounts, lost_accounts) per week."""
    new_accts: dict[str, set[str]] = {}
    lost_accts: dict[str, set[str]] = {}
    for i, wk in enumerate(week_labels):
        if i == 0:
            new_accts[wk] = account_sets[wk]
            lost_accts[wk] = set()
        else:
            prev = account_sets[week_labels[i - 1]]
            curr = account_sets[wk]
            new_accts[wk] = curr - prev
            lost_accts[wk] = prev - curr
    return new_accts, lost_accts


# ---------------------------------------------------------------------------
# 4. Write output workbook
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def write_metric_sheet(ws, title_prefix, dim_name, metric_data, week_labels, fmt=None):
    """Write a single dimension's metric into the worksheet."""
    ws.append([f"{title_prefix} — {dim_name}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(week_labels) + 1)
    ws.cell(1, 1).font = Font(name="Aptos", bold=True, size=13)

    headers = [dim_name] + week_labels
    ws.append(headers)
    style_header(ws, 2, len(headers))

    for value in sorted(metric_data.keys()):
        row_data = [value]
        for v in metric_data[value]:
            if v is None:
                row_data.append("")
            elif fmt == "pct":
                row_data.append(f"{v}%")
            elif fmt == "ratio":
                if v == float("inf"):
                    row_data.append("∞")
                else:
                    row_data.append(f"{v}x")
            elif fmt == "ma":
                row_data.append(round(v, 1) if v is not None else "")
            else:
                row_data.append(v)
        ws.append(row_data)

    auto_width(ws)


def write_account_churn_sheet(ws, new_accts, lost_accts, week_labels):
    ws.append(["Account Churn — New & Disappeared Accounts"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(1, 1).font = Font(name="Aptos", bold=True, size=13)

    headers = ["Week", "New Count", "Lost Count", "Net", "New Accounts", "Lost Accounts"]
    ws.append(headers)
    style_header(ws, 2, len(headers))

    for wk in week_labels:
        new = sorted(new_accts[wk])
        lost = sorted(lost_accts[wk])
        ws.append([
            wk,
            len(new),
            len(lost),
            len(new) - len(lost),
            ", ".join(new) if new else "",
            ", ".join(lost) if lost else "",
        ])

    auto_width(ws)


def write_summary_sheet(ws, data):
    """Highlights: biggest movers, highest momentum, share-of-voice shifts."""
    ws.append(["Trend Summary — Notable Movements"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(1, 1).font = Font(name="Aptos", bold=True, size=13)

    headers = ["Dimension", "Value", "Signal", "Detail", "Direction"]
    ws.append(headers)
    style_header(ws, 2, len(headers))

    week_labels = data["week_labels"]
    if len(week_labels) < 2:
        ws.append(["Not enough weeks for trend analysis"])
        auto_width(ws)
        return

    alerts: list[tuple[str, str, str, str, str]] = []

    for dim_name in DIMENSIONS:
        raw = compute_raw_counts(data["dim_counts"], week_labels, dim_name)
        wow = compute_wow_change(raw)
        sov = compute_share_of_voice(
            data["dim_counts"], week_labels, dim_name, data["week_totals"]
        )
        momentum = compute_momentum(raw)

        for value in raw:
            last_wow = wow[value][-1]
            if last_wow is not None and last_wow != float("inf") and abs(last_wow) >= 50:
                direction = "▲ UP" if last_wow > 0 else "▼ DOWN"
                alerts.append((
                    dim_name, value, "Large WoW Change",
                    f"{last_wow:+.0f}% vs prior week", direction,
                ))

            last_mom = momentum[value][-1] if momentum[value] else None
            if last_mom is not None and last_mom != float("inf"):
                if last_mom >= 1.5:
                    alerts.append((
                        dim_name, value, "High Momentum",
                        f"{last_mom}x (recent {MOVING_AVG_WINDOW}wk avg ÷ prior)", "▲ ACCEL",
                    ))
                elif last_mom <= 0.5:
                    alerts.append((
                        dim_name, value, "Low Momentum",
                        f"{last_mom}x (recent {MOVING_AVG_WINDOW}wk avg ÷ prior)", "▼ DECEL",
                    ))

            sov_vals = sov[value]
            if len(sov_vals) >= 2 and sov_vals[-1] is not None and sov_vals[-2] is not None:
                sov_shift = sov_vals[-1] - sov_vals[-2]
                if abs(sov_shift) >= 5:
                    direction = "▲ UP" if sov_shift > 0 else "▼ DOWN"
                    alerts.append((
                        dim_name, value, "Share-of-Voice Shift",
                        f"{sov_shift:+.1f}pp ({sov_vals[-2]}% → {sov_vals[-1]}%)", direction,
                    ))

    alerts.sort(key=lambda x: x[0])
    for a in alerts:
        ws.append(list(a))

    if not alerts:
        ws.append(["No notable movements detected this period"])

    auto_width(ws)


def build_output(data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    week_labels = data["week_labels"]

    ws_summary = wb.create_sheet("Trend Summary")
    write_summary_sheet(ws_summary, data)

    for dim_name in DIMENSIONS:
        raw = compute_raw_counts(data["dim_counts"], week_labels, dim_name)

        tag = dim_name.replace(" ", "")[:12]

        ws = wb.create_sheet(f"Counts_{tag}")
        write_metric_sheet(ws, "Weekly Counts", dim_name, raw, week_labels)

        ma = compute_moving_average(raw)
        ws = wb.create_sheet(f"MA_{tag}")
        write_metric_sheet(ws, f"{MOVING_AVG_WINDOW}-Week Moving Avg", dim_name, ma, week_labels, fmt="ma")

        wow = compute_wow_change(raw)
        ws = wb.create_sheet(f"WoW_{tag}")
        write_metric_sheet(ws, "Week-over-Week % Change", dim_name, wow, week_labels, fmt="pct")

        sov = compute_share_of_voice(
            data["dim_counts"], week_labels, dim_name, data["week_totals"]
        )
        ws = wb.create_sheet(f"SoV_{tag}")
        write_metric_sheet(ws, "Share of Voice %", dim_name, sov, week_labels, fmt="pct")

        mom = compute_momentum(raw)
        ws = wb.create_sheet(f"Mom_{tag}")
        write_metric_sheet(ws, "Momentum (recent ÷ prior 4wk)", dim_name, mom, week_labels, fmt="ratio")

    new_accts, lost_accts = compute_account_churn(data["account_sets"], week_labels)
    ws = wb.create_sheet("Account Churn")
    write_account_churn_sheet(ws, new_accts, lost_accts, week_labels)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


# ---------------------------------------------------------------------------
# 5. Console summary
# ---------------------------------------------------------------------------

def print_console_summary(data):
    week_labels = data["week_labels"]
    latest = week_labels[-1]
    prev = week_labels[-2] if len(week_labels) >= 2 else None

    print("=" * 70)
    print(f"  Voice of Field — Trend Report")
    print(f"  Latest week: {latest}  |  {data['week_totals'][latest]} insights")
    print(f"  Weeks analyzed: {len(week_labels)} ({week_labels[0]} → {latest})")
    print("=" * 70)

    if prev:
        total_change = data["week_totals"][latest] - data["week_totals"][prev]
        sign = "+" if total_change >= 0 else ""
        print(f"\n  Total insights change vs prior week: {sign}{total_change}"
              f" ({data['week_totals'][prev]} → {data['week_totals'][latest]})")

    for dim_name in DIMENSIONS:
        raw = compute_raw_counts(data["dim_counts"], week_labels, dim_name)
        ma = compute_moving_average(raw)
        print(f"\n  ─── {dim_name} ({MOVING_AVG_WINDOW}-wk MA, latest) ───")
        for value in sorted(raw.keys()):
            latest_count = raw[value][-1]
            latest_ma = ma[value][-1]
            ma_str = f"{latest_ma:.1f}" if latest_ma is not None else "n/a"
            print(f"    {value:<40s}  count={latest_count:>3d}   MA={ma_str:>6s}")

    new_accts, lost_accts = compute_account_churn(data["account_sets"], week_labels)
    print(f"\n  ─── Account Churn (latest week) ───")
    print(f"    New:  {len(new_accts[latest]):>3d}  |  Lost: {len(lost_accts[latest]):>3d}"
          f"  |  Net: {len(new_accts[latest]) - len(lost_accts[latest]):+d}")
    if new_accts[latest]:
        print(f"    New accounts: {', '.join(sorted(new_accts[latest])[:10])}")

    print(f"\n  Output saved to: {OUTPUT_PATH}")
    print("=" * 70)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    data = load_weekly_data(INPUT_PATH)
    out = build_output(data)
    print_console_summary(data)
