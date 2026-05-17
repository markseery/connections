"""Analyze imported Robinhood JSON records and compute current positions."""

from __future__ import annotations

import json
import random
import re
import statistics
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from common.simple.yfinance_warnings import suppress_utcnow_deprecation_warning
import yfinance as yf
import yaml
from openpyxl import Workbook

from common.compound.finance_pipeline.distribution_history_comparison import DistributionHistoryComparison
from common.compound.finance_pipeline.price_history_growth import PriceHistoryGrowthAnalyzer
from common.compound.finance_pipeline.stockanalysis_dividends import StockAnalysisDividendExtractor

suppress_utcnow_deprecation_warning()


BUY_LIKE_CODES = {"BUY", "ACATI"}
SELL_LIKE_CODES = {"SELL", "ACATO"}
DIVIDEND_CODES = {"CDIV", "MDIV"}
OPTIONS_CODES = {"STO", "BTC"}
MONTHLY_POSITION_CHANGE_CODES = BUY_LIKE_CODES | SELL_LIKE_CODES
_DIV_SHARES_AT = re.compile(
    r"([\d,]+(?:\.\d+)?)\s+shares\s+at\s+([\d,]+(?:\.\d+)?)",
    re.I,
)


@dataclass
class PositionSummary:
    account_name: str
    records_read: int
    positions: list[dict[str, Any]]
    cash_reconciliation: dict[str, Any]
    cash_forecast: dict[str, Any]


class RobinhoodPositionAnalyzer:
    """Build a current-positions summary from imported per-record JSON files."""

    def __init__(
        self,
        account_name: str,
        portfolio_dir: Path | None = None,
        monte_carlo_constraints_path: Path | None = None,
    ) -> None:
        self.account_name = account_name.strip()
        if not self.account_name:
            raise ValueError("account_name must not be empty")
        root = portfolio_dir or Path("application_files/data/portfolio")
        self.portfolio_dir = root.resolve()
        self.account_dir = (self.portfolio_dir / "robinhood" / self.account_name).resolve()
        self._price_cache: dict[str, float | None] = {}
        self._historical_price_cache: dict[tuple[str, str], float | None] = {}
        self._dividend_yield_cache: dict[str, float | None] = {}
        self._stockanalysis_yield_cache: dict[str, float | None] = {}
        self._stockanalysis_extractor = StockAnalysisDividendExtractor()
        self.monte_carlo_constraints_path = (
            monte_carlo_constraints_path.resolve()
            if monte_carlo_constraints_path is not None
            else Path("application_files/config/portfolio_monte_carlo_constraints.yaml").resolve()
        )
        self.excluded_symbols = self._load_excluded_symbols()
        self.monte_carlo_constraints = self._load_monte_carlo_constraints()

    def _load_excluded_symbols(self) -> set[str]:
        config_path = Path("application_files/config/portfolio_symbol_filters.yaml").resolve()
        if not config_path.exists():
            return set()
        try:
            parsed = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
            values = parsed.get("exclude_symbols") or []
            if not isinstance(values, list):
                return set()
            out: set[str] = set()
            for item in values:
                sym = str(item or "").strip().upper()
                if sym:
                    out.add(sym)
            return out
        except Exception:
            return set()

    @staticmethod
    def _parse_pct(value: Any) -> float | None:
        parsed = RobinhoodPositionAnalyzer._to_float(value)
        if parsed is None:
            return None
        if parsed <= 0:
            return None
        return min(100.0, float(parsed))

    @staticmethod
    def _parse_nonnegative_float(value: Any) -> float | None:
        parsed = RobinhoodPositionAnalyzer._to_float(value)
        if parsed is None:
            return None
        if parsed < 0:
            return None
        return float(parsed)

    @staticmethod
    def _parse_positive_int(value: Any) -> int | None:
        try:
            if value is None:
                return None
            parsed = int(value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    def _load_monte_carlo_constraints(self) -> dict[str, Any]:
        config_path = self.monte_carlo_constraints_path
        default: dict[str, Any] = {
            "simulation_count": None,
            "sampling_alpha_scale": 1.0,
            "include_symbols_even_if_not_held": [],
            "max_symbol_weight_pct": None,
            "max_top_5_weight_pct": None,
            "max_top_10_weight_pct": None,
            "confidence_tier_symbol_caps": [],
            "high_confidence_symbol_cap": None,
            "max_growth_contribution_pct": None,
            "max_income_contribution_pct": None,
            "theme_caps": [],
            "allocation_confidence_growth_weight": 0.45,
            "allocation_confidence_dividend_weight": 0.55,
            "allocation_confidence_multiplier_min": 0.8,
            "allocation_confidence_multiplier_max": 1.2,
            "min_accepted_run_pct_target": None,
            "min_price_history_months": None,
            "min_confidence_score": None,
            "low_confidence_score_threshold": None,
            "max_low_confidence_total_weight_pct": None,
            "max_annualized_growth_pct_for_projection": None,
            "scenario_constraints": {},
            "top_concentration_shocks_pct": [20.0, 30.0],
        }
        if not config_path.exists():
            return default
        try:
            parsed = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
            if not isinstance(parsed, dict):
                return default

            def _nested(*keys: str) -> Any:
                cur: Any = parsed
                for key in keys:
                    if not isinstance(cur, dict):
                        return None
                    cur = cur.get(key)
                return cur

            def _pick_pct(*candidates: Any) -> float | None:
                for c in candidates:
                    v = self._parse_pct(c)
                    if v is not None:
                        return v
                return None

            out = dict(default)
            out["simulation_count"] = self._parse_positive_int(
                parsed.get("simulation_count")
            ) or self._parse_positive_int(_nested("simulation", "simulation_count"))
            include_symbols = (
                _nested("simulation", "include_symbols_even_if_not_held")
                or parsed.get("include_symbols_even_if_not_held")
                or []
            )
            if isinstance(include_symbols, list):
                out["include_symbols_even_if_not_held"] = sorted(
                    {
                        str(s or "").strip().upper()
                        for s in include_symbols
                        if str(s or "").strip()
                    }
                )
            sampling_alpha_scale = self._parse_nonnegative_float(
                _nested("simulation", "sampling_alpha_scale")
            )
            if sampling_alpha_scale is not None and sampling_alpha_scale > 0:
                out["sampling_alpha_scale"] = float(sampling_alpha_scale)
            out["max_symbol_weight_pct"] = _pick_pct(
                parsed.get("max_symbol_weight_pct"),
                _nested("allocation_constraints", "symbol", "max_weight_pct"),
            )
            out["max_top_5_weight_pct"] = _pick_pct(
                parsed.get("max_top_5_weight_pct"),
                _nested("allocation_constraints", "symbol", "max_top_5_weight_pct"),
            )
            out["max_top_10_weight_pct"] = _pick_pct(
                parsed.get("max_top_10_weight_pct"),
                _nested("allocation_constraints", "symbol", "max_top_10_weight_pct"),
            )
            raw_tiers = (
                parsed.get("confidence_tier_symbol_caps")
                or _nested("allocation_constraints", "symbol", "confidence_tier_caps")
                or []
            )
            tiers: list[dict[str, float]] = []
            if isinstance(raw_tiers, list):
                for row in raw_tiers:
                    if not isinstance(row, dict):
                        continue
                    max_conf = self._to_float(row.get("max_confidence_score"))
                    max_weight_pct = self._parse_pct(row.get("max_weight_pct"))
                    if (
                        max_conf is None
                        or max_weight_pct is None
                        or max_conf < 0
                        or max_conf > 100
                    ):
                        continue
                    tiers.append(
                        {
                            "max_confidence_score": float(max_conf),
                            "max_weight_pct": float(max_weight_pct),
                        }
                    )
            tiers.sort(key=lambda r: r["max_confidence_score"])
            out["confidence_tier_symbol_caps"] = tiers
            high_conf_row = (
                _nested("allocation_constraints", "symbol", "high_confidence_cap")
                or parsed.get("high_confidence_symbol_cap")
            )
            if isinstance(high_conf_row, dict):
                min_conf = self._to_float(high_conf_row.get("min_confidence_score"))
                max_weight_pct = self._parse_pct(high_conf_row.get("max_weight_pct"))
                if (
                    min_conf is not None
                    and max_weight_pct is not None
                    and 0 <= min_conf <= 100
                ):
                    out["high_confidence_symbol_cap"] = {
                        "min_confidence_score": float(min_conf),
                        "max_weight_pct": float(max_weight_pct),
                    }
            out["max_growth_contribution_pct"] = _pick_pct(
                parsed.get("max_growth_contribution_pct"),
                _nested("contribution_constraints", "max_growth_contribution_pct"),
            )
            out["max_income_contribution_pct"] = _pick_pct(
                parsed.get("max_income_contribution_pct"),
                _nested("contribution_constraints", "max_income_contribution_pct"),
            )

            growth_w = self._to_float(
                parsed.get("allocation_confidence_growth_weight")
            )
            if growth_w is None:
                growth_w = self._to_float(
                    _nested("confidence_controls", "allocation_confidence", "growth_weight")
                )
            dividend_w = self._to_float(
                parsed.get("allocation_confidence_dividend_weight")
            )
            if dividend_w is None:
                dividend_w = self._to_float(
                    _nested("confidence_controls", "allocation_confidence", "dividend_weight")
                )
            if growth_w is not None and dividend_w is not None and (growth_w + dividend_w) > 0:
                s = float(growth_w + dividend_w)
                out["allocation_confidence_growth_weight"] = float(growth_w) / s
                out["allocation_confidence_dividend_weight"] = float(dividend_w) / s

            min_mult = self._to_float(parsed.get("allocation_confidence_multiplier_min"))
            if min_mult is None:
                min_mult = self._to_float(
                    _nested("confidence_controls", "allocation_confidence", "multiplier_min")
                )
            max_mult = self._to_float(parsed.get("allocation_confidence_multiplier_max"))
            if max_mult is None:
                max_mult = self._to_float(
                    _nested("confidence_controls", "allocation_confidence", "multiplier_max")
                )
            if (
                min_mult is not None
                and max_mult is not None
                and min_mult > 0
                and max_mult > 0
                and min_mult <= max_mult
            ):
                out["allocation_confidence_multiplier_min"] = float(min_mult)
                out["allocation_confidence_multiplier_max"] = float(max_mult)

            out["min_accepted_run_pct_target"] = _pick_pct(
                parsed.get("min_accepted_run_pct_target"),
                _nested("simulation", "min_accepted_run_pct_target"),
            )
            out["min_price_history_months"] = self._parse_nonnegative_float(
                _nested("data_quality_constraints", "min_price_history_months")
            )
            out["min_confidence_score"] = _pick_pct(
                _nested("data_quality_constraints", "min_confidence_score")
            )
            out["low_confidence_score_threshold"] = _pick_pct(
                _nested("data_quality_constraints", "low_confidence_score_threshold")
            )
            out["max_low_confidence_total_weight_pct"] = _pick_pct(
                _nested("data_quality_constraints", "max_low_confidence_total_weight_pct")
            )
            out["max_annualized_growth_pct_for_projection"] = _pick_pct(
                _nested("growth_projection_controls", "max_annualized_growth_pct_for_projection")
            )

            theme_caps: list[dict[str, Any]] = []
            raw_theme_caps = parsed.get("theme_caps")
            if raw_theme_caps is None:
                raw_theme_caps = _nested("allocation_constraints", "theme_caps")
            raw_theme_caps = raw_theme_caps or []
            if isinstance(raw_theme_caps, list):
                for row in raw_theme_caps:
                    if not isinstance(row, dict):
                        continue
                    max_weight_pct = _pick_pct(
                        row.get("max_weight_pct"),
                        row.get("max_theme_weight_pct"),
                    )
                    if max_weight_pct is None:
                        continue
                    symbols_raw = row.get("symbols") or []
                    if not isinstance(symbols_raw, list):
                        continue
                    symbols = sorted(
                        {
                            str(s or "").strip().upper()
                            for s in symbols_raw
                            if str(s or "").strip()
                        }
                    )
                    if not symbols:
                        continue
                    theme_caps.append(
                        {
                            "name": str(row.get("name") or "theme").strip() or "theme",
                            "max_weight_pct": max_weight_pct,
                            "symbols": symbols,
                        }
                    )
            out["theme_caps"] = theme_caps
            scenario_constraints: dict[str, Any] = {}
            raw_scenarios = _nested("scenario_constraints") or {}
            if isinstance(raw_scenarios, dict):
                for scenario_name, row in raw_scenarios.items():
                    if not isinstance(row, dict):
                        continue
                    scenario_constraints[str(scenario_name)] = {
                        "min_income_30_day_distribution": self._parse_nonnegative_float(
                            row.get("min_income_30_day_distribution")
                        ),
                        "min_growth_30_day_dollars": self._parse_nonnegative_float(
                            row.get("min_growth_30_day_dollars")
                        ),
                        "max_low_confidence_weight_pct": _pick_pct(
                            row.get("max_low_confidence_weight_pct")
                        ),
                    }
            out["scenario_constraints"] = scenario_constraints
            raw_shocks = (
                _nested("risk_controls", "top_concentration_shocks_pct")
                or _nested("stress_tests", "top_concentration_shocks_pct")
                or []
            )
            if isinstance(raw_shocks, list):
                cleaned: list[float] = []
                for s in raw_shocks:
                    fv = self._parse_nonnegative_float(s)
                    if fv is None or fv <= 0:
                        continue
                    cleaned.append(float(fv))
                if cleaned:
                    out["top_concentration_shocks_pct"] = sorted(set(cleaned))
            return out
        except Exception:
            return default

    def set_monte_carlo_overrides(self, overrides: dict[str, Any]) -> None:
        if not overrides:
            return
        merged = dict(self.monte_carlo_constraints or {})
        for key, value in overrides.items():
            if value is None:
                continue
            merged[key] = value
        self.monte_carlo_constraints = merged

    @staticmethod
    def _normalize_weights(values: list[float]) -> list[float]:
        clean = [max(0.0, float(v)) for v in values]
        total = sum(clean)
        if total <= 0:
            n = len(clean)
            return ([1.0 / n] * n) if n > 0 else []
        return [v / total for v in clean]

    @staticmethod
    def _apply_symbol_weight_caps(weights: list[float], caps: list[float]) -> list[float]:
        n = len(weights)
        if n == 0:
            return []
        w = RobinhoodPositionAnalyzer._normalize_weights(weights)
        limits = [max(0.0, min(1.0, float(c))) for c in caps]
        sum_limits = sum(limits)
        if sum_limits <= 0:
            return w
        if sum_limits < 1.0:
            # Relax infeasible caps proportionally to allow a valid 100% allocation.
            relax = 1.0 / sum_limits
            limits = [max(0.0, min(1.0, c * relax)) for c in limits]

        fixed = [False] * n
        for _ in range(n + 2):
            changed = False
            for i in range(n):
                if fixed[i]:
                    continue
                if w[i] > limits[i]:
                    w[i] = limits[i]
                    fixed[i] = True
                    changed = True
            if not changed:
                break
            remaining = 1.0 - sum(w[i] for i in range(n) if fixed[i])
            free_idx = [i for i in range(n) if not fixed[i]]
            if not free_idx:
                break
            base = sum(w[i] for i in free_idx)
            if base <= 0:
                even = remaining / len(free_idx)
                for i in free_idx:
                    w[i] = even
            else:
                for i in free_idx:
                    w[i] = remaining * (w[i] / base)
        return RobinhoodPositionAnalyzer._normalize_weights(w)

    @staticmethod
    def _apply_theme_caps(
        weights: list[float],
        theme_caps: list[tuple[set[int], float]],
    ) -> list[float]:
        if not weights or not theme_caps:
            return RobinhoodPositionAnalyzer._normalize_weights(weights)
        w = RobinhoodPositionAnalyzer._normalize_weights(weights)
        n = len(w)
        for _ in range(12):
            changed = False
            for member_idx, cap in theme_caps:
                cap = max(0.0, min(1.0, cap))
                idx = sorted(i for i in member_idx if 0 <= i < n)
                if not idx:
                    continue
                group_weight = sum(w[i] for i in idx)
                if group_weight <= cap + 1e-12:
                    continue
                changed = True
                excess = group_weight - cap
                shrink = cap / group_weight if group_weight > 0 else 0.0
                for i in idx:
                    w[i] *= shrink
                outside = [i for i in range(n) if i not in member_idx]
                if not outside:
                    continue
                outside_total = sum(w[i] for i in outside)
                if outside_total <= 0:
                    add = excess / len(outside)
                    for i in outside:
                        w[i] += add
                else:
                    for i in outside:
                        w[i] += excess * (w[i] / outside_total)
            if not changed:
                break
        return RobinhoodPositionAnalyzer._normalize_weights(w)

    @staticmethod
    def _read_json(path: Path) -> dict[str, Any] | None:
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return None

    @staticmethod
    def _to_float(value: Any) -> float | None:
        if value is None:
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _get_current_price(self, symbol: str) -> float | None:
        sym = symbol.strip().upper()
        if not sym:
            return None
        if sym in self._price_cache:
            return self._price_cache[sym]
        try:
            ticker = yf.Ticker(sym)
            info = ticker.info or {}
            current_price = info.get("regularMarketPrice") or info.get("currentPrice")
            if current_price is None:
                hist = ticker.history(period="2d")
                if hist is not None and not hist.empty:
                    current_price = float(hist.iloc[-1]["Close"])
            value = self._to_float(current_price)
            self._price_cache[sym] = value
            return value
        except Exception:
            self._price_cache[sym] = None
            return None

    def _get_historical_close(self, symbol: str, date_iso: Any) -> float | None:
        sym = symbol.strip().upper()
        d = str(date_iso or "").strip()
        if not sym or not d:
            return None
        key = (sym, d)
        if key in self._historical_price_cache:
            return self._historical_price_cache[key]
        try:
            day = datetime.strptime(d, "%Y-%m-%d").date()
            ticker = yf.Ticker(sym)
            # Pull a short window to handle weekends/holidays and pick first available close.
            hist = ticker.history(
                start=day.isoformat(),
                end=(day + timedelta(days=7)).isoformat(),
                auto_adjust=True,
            )
            if hist is None or hist.empty:
                self._historical_price_cache[key] = None
                return None
            close = self._to_float(hist.iloc[0].get("Close"))
            self._historical_price_cache[key] = close
            return close
        except Exception:
            self._historical_price_cache[key] = None
            return None

    def _get_dividend_yield(self, symbol: str) -> float | None:
        """Return annual dividend yield as decimal fraction (e.g. 0.04 == 4%)."""
        sym = symbol.strip().upper()
        if not sym:
            return None
        if sym in self._dividend_yield_cache:
            return self._dividend_yield_cache[sym]
        try:
            ticker = yf.Ticker(sym)
            info = ticker.info or {}
            candidate = self._to_float(info.get("dividendYield"))
            if candidate is None:
                candidate = self._to_float(info.get("trailingAnnualDividendYield"))
            if candidate is None:
                annual_rate = self._to_float(info.get("trailingAnnualDividendRate"))
                price = self._get_current_price(sym)
                if annual_rate is not None and price and price > 0:
                    candidate = annual_rate / price
            if candidate is None:
                self._dividend_yield_cache[sym] = None
                return None
            # Normalize occasional percent-style values (e.g. 4.2 means 4.2%).
            if candidate > 1:
                candidate = candidate / 100.0
            if candidate <= 0:
                self._dividend_yield_cache[sym] = None
                return None
            self._dividend_yield_cache[sym] = candidate
            return candidate
        except Exception:
            self._dividend_yield_cache[sym] = None
            return None

    @staticmethod
    def _median_day_gap_from_iso_dates(date_values: list[str]) -> float | None:
        parsed: list[date] = []
        for raw in date_values:
            d = RobinhoodPositionAnalyzer._parse_iso_date(raw)
            if d is not None:
                parsed.append(d)
        unique_sorted = sorted(set(parsed))
        if len(unique_sorted) < 2:
            return None
        gaps = [
            (curr - prev).days
            for prev, curr in zip(unique_sorted, unique_sorted[1:])
            if (curr - prev).days > 0
        ]
        if not gaps:
            return None
        return float(statistics.median(gaps))

    def _get_stockanalysis_yield(self, symbol: str) -> float | None:
        """Estimate annual dividend yield as decimal using StockAnalysis data."""
        sym = symbol.strip().upper()
        if not sym:
            return None
        if sym in self._stockanalysis_yield_cache:
            return self._stockanalysis_yield_cache[sym]
        try:
            # Prefer annual yield from ETF overview page (Dividend Yield field).
            overview_yield = self._stockanalysis_extractor.extract_overview_annual_yield(sym)
            if overview_yield is not None and overview_yield > 0:
                self._stockanalysis_yield_cache[sym] = overview_yield
                return overview_yield

            snapshot = self._stockanalysis_extractor.extract(sym)
            direct_pct = self._to_float(snapshot.dividend_yield_pct)
            if direct_pct is not None and direct_pct > 0:
                y = direct_pct / 100.0
                self._stockanalysis_yield_cache[sym] = y
                return y
            annual_dividend = self._to_float(snapshot.annual_dividend)
            current_price = self._get_current_price(sym)
            if (
                annual_dividend is not None
                and annual_dividend > 0
                and current_price is not None
                and current_price > 0
            ):
                y = annual_dividend / current_price
                self._stockanalysis_yield_cache[sym] = y
                return y
            cash_amounts = [
                float(row.cash_amount)
                for row in snapshot.history
                if row.cash_amount is not None and float(row.cash_amount) > 0
            ]
            if not cash_amounts:
                self._stockanalysis_yield_cache[sym] = None
                return None
            median_cash = statistics.median(cash_amounts)
            median_gap = self._median_day_gap_from_iso_dates(
                [str(row.ex_dividend_date or "") for row in snapshot.history]
            )
            if median_gap is None:
                self._stockanalysis_yield_cache[sym] = None
                return None
            current_price = self._get_current_price(sym)
            if current_price is None or current_price <= 0:
                self._stockanalysis_yield_cache[sym] = None
                return None
            annual_rate = median_cash * (365.0 / median_gap)
            if annual_rate <= 0:
                self._stockanalysis_yield_cache[sym] = None
                return None
            y = annual_rate / current_price
            self._stockanalysis_yield_cache[sym] = y
            return y
        except Exception:
            self._stockanalysis_yield_cache[sym] = None
            return None

    @staticmethod
    def _sort_date(value: Any) -> str:
        s = str(value or "").strip()
        return s if s else "9999-12-31"

    @staticmethod
    def _sort_row_number(value: Any) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    def _record_sort_key(self, rec: dict[str, Any]) -> tuple[str, str, str, int, int]:
        trans_code = str(rec.get("trans_code") or "").strip().upper()
        if trans_code in BUY_LIKE_CODES:
            trans_priority = 0
        elif trans_code in SELL_LIKE_CODES:
            trans_priority = 1
        else:
            trans_priority = 2
        return (
            self._sort_date(rec.get("activity_date")),
            self._sort_date(rec.get("process_date")),
            self._sort_date(rec.get("settle_date")),
            trans_priority,
            self._sort_row_number(rec.get("row_number")),
        )

    @staticmethod
    def _extract_div_qty_price(description: Any) -> tuple[float | None, float | None]:
        text = str(description or "")
        m = _DIV_SHARES_AT.search(text)
        if not m:
            return None, None
        try:
            qty = float(m.group(1).replace(",", ""))
            price = float(m.group(2).replace(",", ""))
            return qty, price
        except (TypeError, ValueError):
            return None, None

    @staticmethod
    def _is_weekly_dividend_pattern(dividend_dates: list[str]) -> bool:
        parsed_dates: list[datetime] = []
        for value in dividend_dates:
            text = str(value or "").strip()
            if not text:
                continue
            try:
                parsed_dates.append(datetime.strptime(text, "%Y-%m-%d"))
            except ValueError:
                continue
        if len(parsed_dates) < 2:
            return False
        unique_sorted = sorted({d.date() for d in parsed_dates})
        if len(unique_sorted) < 2:
            return False
        day_gaps = [
            (curr - prev).days
            for prev, curr in zip(unique_sorted, unique_sorted[1:])
            if (curr - prev).days > 0
        ]
        if not day_gaps:
            return False
        weekly_like = [gap for gap in day_gaps if 5 <= gap <= 9]
        if not weekly_like:
            return False
        median_gap = statistics.median(day_gaps)
        return 5 <= median_gap <= 9 and len(weekly_like) >= (len(day_gaps) / 2)

    @staticmethod
    def _parse_iso_date(value: Any) -> date | None:
        text = str(value or "").strip()
        if not text:
            return None
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            return None

    @staticmethod
    def _cadence_from_median_gap(median_gap: float) -> tuple[int, str] | None:
        if median_gap <= 0:
            return None
        if median_gap <= 10:
            return 7, "weekly"
        if median_gap <= 20:
            return 14, "biweekly"
        if median_gap <= 45:
            return 30, "monthly"
        if median_gap <= 120:
            return 91, "quarterly"
        return None

    def _build_cash_forecast(
        self,
        dividend_events_by_symbol: dict[str, list[dict[str, Any]]],
        current_position_qty_by_symbol: dict[str, float],
        position_snapshot_by_symbol: dict[str, dict[str, Any]],
        as_of_date: date,
    ) -> dict[str, Any]:
        horizons = (30, 60, 90)
        horizon_end = as_of_date + timedelta(days=max(horizons))
        symbols: list[dict[str, Any]] = []
        total_30 = 0.0
        total_60 = 0.0
        total_90 = 0.0

        def _parse_any_date(text: str | None) -> date | None:
            raw = str(text or "").strip()
            if not raw:
                return None
            for fmt in ("%Y-%m-%d", "%b %d, %Y", "%B %d, %Y"):
                try:
                    return datetime.strptime(raw, fmt).date()
                except ValueError:
                    continue
            return None

        def _build_projected_dates_and_totals(
            last_date: date, cadence_days: int, cash_amount: float
        ) -> tuple[list[date], float, float, float]:
            projected_dates: list[date] = []
            projected_30 = 0.0
            projected_60 = 0.0
            projected_90 = 0.0
            next_date = last_date
            while next_date <= horizon_end:
                next_date = next_date + timedelta(days=cadence_days)
                if next_date <= as_of_date:
                    continue
                if next_date > horizon_end:
                    break
                projected_dates.append(next_date)
                if next_date <= as_of_date + timedelta(days=30):
                    projected_30 += cash_amount
                if next_date <= as_of_date + timedelta(days=60):
                    projected_60 += cash_amount
                if next_date <= as_of_date + timedelta(days=90):
                    projected_90 += cash_amount
            return projected_dates, projected_30, projected_60, projected_90

        def _stockanalysis_history_based_projection(
            symbol: str,
            qty: float,
            base_note: str,
            events_observed: int = 0,
            last_dividend_date: str | None = None,
            median_dividend_amount: float | None = None,
        ) -> dict[str, Any] | None:
            try:
                snapshot = self._stockanalysis_extractor.extract(symbol)
            except Exception:
                return None
            amounts = [
                float(row.cash_amount)
                for row in snapshot.history
                if row.cash_amount is not None and float(row.cash_amount) > 0
            ]
            if not amounts:
                return None
            parsed_dates = [
                _parse_any_date(row.ex_dividend_date) for row in snapshot.history
            ]
            parsed_dates = [d for d in parsed_dates if d is not None]
            unique_dates = sorted(set(parsed_dates))
            cadence_days: int | None = None
            cadence_label: str | None = None
            if len(unique_dates) >= 2:
                gaps = [
                    (curr - prev).days
                    for prev, curr in zip(unique_dates, unique_dates[1:])
                    if (curr - prev).days > 0
                ]
                if gaps:
                    cadence_info = self._cadence_from_median_gap(statistics.median(gaps))
                    if cadence_info is not None:
                        cadence_days, cadence_label = cadence_info
            if cadence_days is None:
                freq = str(snapshot.payout_frequency or "").strip().lower()
                if "weekly" in freq:
                    cadence_days, cadence_label = 7, "weekly"
                elif "biweekly" in freq:
                    cadence_days, cadence_label = 14, "biweekly"
                elif "month" in freq:
                    cadence_days, cadence_label = 30, "monthly"
                elif "quarter" in freq:
                    cadence_days, cadence_label = 91, "quarterly"
            if cadence_days is None:
                return None
            median_amt_per_share = statistics.median(amounts)
            if median_amt_per_share <= 0:
                return None
            median_amt_total = median_amt_per_share * qty
            if median_amt_total <= 0:
                return None
            last_date = unique_dates[-1] if unique_dates else _parse_any_date(
                snapshot.ex_dividend_date
            )
            if last_date is None:
                return None
            projected_dates, projected_30, projected_60, projected_90 = (
                _build_projected_dates_and_totals(
                    last_date=last_date,
                    cadence_days=cadence_days,
                    cash_amount=median_amt_total,
                )
            )
            if not projected_dates:
                return None
            return {
                "instrument": symbol,
                "cadence_label": cadence_label,
                "cadence_days": cadence_days,
                "events_observed": max(events_observed, len(snapshot.history)),
                "last_dividend_date": last_date.isoformat(),
                "median_dividend_amount": round(median_amt_total, 2),
                "projected_next_30_days": round(projected_30, 2),
                "projected_next_60_days": round(projected_60, 2),
                "projected_next_90_days": round(projected_90, 2),
                "projected_payout_dates": [d.isoformat() for d in projected_dates],
                "forecast_note": f"{base_note}_stockanalysis_history_fallback",
                "yield_source": "stockanalysis_history",
                "dividend_yield_used": None,
                "per_share_distribution_rate_used": round(median_amt_per_share, 6),
                "shares_used_for_projection": round(qty, 6),
                "current_value": round(pos_current_value, 2) if pos_current_value is not None else None,
                "estimated_cost_basis": (
                    round(pos_estimated_cost_basis, 2)
                    if pos_estimated_cost_basis is not None
                    else None
                ),
                "total_pl_pct_of_cost_basis": (
                    round(pos_total_pl_pct, 4) if pos_total_pl_pct is not None else None
                ),
            }

        def _yield_based_projection(
            symbol: str,
            qty: float,
            base_note: str,
            events_observed: int = 0,
            last_dividend_date: str | None = None,
            median_dividend_amount: float | None = None,
        ) -> dict[str, Any]:
            history_row = _stockanalysis_history_based_projection(
                symbol=symbol,
                qty=qty,
                base_note=base_note,
                events_observed=events_observed,
                last_dividend_date=last_dividend_date,
                median_dividend_amount=median_dividend_amount,
            )
            if history_row is not None:
                return history_row
            price = self._get_current_price(symbol)
            div_yield = self._get_dividend_yield(symbol)
            yield_source = "yfinance"
            if div_yield is None or div_yield <= 0:
                div_yield = self._get_stockanalysis_yield(symbol)
                yield_source = "stockanalysis"
            if price is None or price <= 0 or div_yield is None or div_yield <= 0:
                return {
                    "instrument": symbol,
                    "cadence_label": None,
                    "cadence_days": None,
                    "events_observed": events_observed,
                    "last_dividend_date": last_dividend_date,
                    "median_dividend_amount": median_dividend_amount,
                    "projected_next_30_days": 0.0,
                    "projected_next_60_days": 0.0,
                    "projected_next_90_days": 0.0,
                    "projected_payout_dates": [],
                    "forecast_note": f"{base_note}_and_yield_unavailable",
                    "yield_source": None,
                    "dividend_yield_used": None,
                    "per_share_distribution_rate_used": None,
                    "shares_used_for_projection": round(qty, 6),
                    "current_value": (
                        round(pos_current_value, 2) if pos_current_value is not None else None
                    ),
                    "estimated_cost_basis": (
                        round(pos_estimated_cost_basis, 2)
                        if pos_estimated_cost_basis is not None
                        else None
                    ),
                    "total_pl_pct_of_cost_basis": (
                        round(pos_total_pl_pct, 4) if pos_total_pl_pct is not None else None
                    ),
                }
            annual_income = qty * price * div_yield
            projected_30 = annual_income * (30.0 / 365.0)
            projected_60 = annual_income * (60.0 / 365.0)
            projected_90 = annual_income * (90.0 / 365.0)
            return {
                "instrument": symbol,
                "cadence_label": None,
                "cadence_days": None,
                "events_observed": events_observed,
                "last_dividend_date": last_dividend_date,
                "median_dividend_amount": median_dividend_amount,
                "projected_next_30_days": round(projected_30, 2),
                "projected_next_60_days": round(projected_60, 2),
                "projected_next_90_days": round(projected_90, 2),
                "projected_payout_dates": [],
                "forecast_note": f"{base_note}_yield_fallback",
                "yield_source": yield_source,
                "dividend_yield_used": round(div_yield, 6),
                # Annual yield estimate converted to per-share 30-day amount.
                "per_share_distribution_rate_used": round((price * div_yield) * (30.0 / 365.0), 6),
                "shares_used_for_projection": round(qty, 6),
                "current_value": round(pos_current_value, 2) if pos_current_value is not None else None,
                "estimated_cost_basis": (
                    round(pos_estimated_cost_basis, 2)
                    if pos_estimated_cost_basis is not None
                    else None
                ),
                "total_pl_pct_of_cost_basis": (
                    round(pos_total_pl_pct, 4) if pos_total_pl_pct is not None else None
                ),
            }

        for symbol in sorted(current_position_qty_by_symbol.keys()):
            current_qty = float(current_position_qty_by_symbol.get(symbol) or 0.0)
            if current_qty <= 0:
                continue
            pos = position_snapshot_by_symbol.get(symbol, {})
            buy_qty = self._to_float(pos.get("buy_quantity")) or 0.0
            # Ignore tiny residual stubs (e.g. 1-2 shares left from large round-trip trades)
            # so they do not distort rebalancing optimization output.
            if (
                current_qty <= 2.0
                and buy_qty > 0
                and (current_qty / buy_qty) <= 0.01
            ):
                continue
            pos_current_value = self._to_float(pos.get("current_value"))
            pos_estimated_cost_basis = self._to_float(pos.get("estimated_cost_basis"))
            pos_total_pl_pct = self._to_float(pos.get("total_pl_pct_of_cost_basis"))
            events = dividend_events_by_symbol.get(symbol, [])
            if not events:
                row = _yield_based_projection(
                    symbol=symbol,
                    qty=current_qty,
                    base_note="no_dividend_history_for_open_position",
                )
                symbols.append(row)
                total_30 += float(row.get("projected_next_30_days") or 0.0)
                total_60 += float(row.get("projected_next_60_days") or 0.0)
                total_90 += float(row.get("projected_next_90_days") or 0.0)
                continue
            events = sorted(events, key=lambda x: x.get("date") or date.min)
            amounts = [
                float(e.get("amount_total") or 0.0)
                for e in events
                if float(e.get("amount_total") or 0.0) > 0
            ]
            per_share_rates = [
                float(e.get("amount_per_share") or 0.0)
                for e in events
                if float(e.get("amount_per_share") or 0.0) > 0
            ]
            if len(amounts) < 2:
                median_rate = (
                    statistics.median(per_share_rates) if per_share_rates else None
                )
                row = _yield_based_projection(
                    symbol=symbol,
                    qty=current_qty,
                    base_note="insufficient_dividend_history",
                    events_observed=len(events),
                    last_dividend_date=(events[-1].get("date") or date.min).isoformat()
                    if events
                    else None,
                    median_dividend_amount=(
                        round(median_rate * current_qty, 2)
                        if median_rate is not None
                        else (round(statistics.median(amounts), 2) if amounts else None)
                    ),
                )
                symbols.append(row)
                total_30 += float(row.get("projected_next_30_days") or 0.0)
                total_60 += float(row.get("projected_next_60_days") or 0.0)
                total_90 += float(row.get("projected_next_90_days") or 0.0)
                continue
            unique_dates = sorted(
                {
                    e.get("date")
                    for e in events
                    if isinstance(e.get("date"), date)
                }
            )
            if len(unique_dates) < 2:
                median_rate = (
                    statistics.median(per_share_rates) if per_share_rates else None
                )
                row = _yield_based_projection(
                    symbol=symbol,
                    qty=current_qty,
                    base_note="insufficient_unique_dividend_dates",
                    events_observed=len(events),
                    last_dividend_date=unique_dates[-1].isoformat(),
                    median_dividend_amount=(
                        round(median_rate * current_qty, 2)
                        if median_rate is not None
                        else round(statistics.median(amounts), 2)
                    ),
                )
                symbols.append(row)
                total_30 += float(row.get("projected_next_30_days") or 0.0)
                total_60 += float(row.get("projected_next_60_days") or 0.0)
                total_90 += float(row.get("projected_next_90_days") or 0.0)
                continue
            day_gaps = [
                (curr - prev).days
                for prev, curr in zip(unique_dates, unique_dates[1:])
                if (curr - prev).days > 0
            ]
            if not day_gaps:
                median_rate = (
                    statistics.median(per_share_rates) if per_share_rates else None
                )
                row = _yield_based_projection(
                    symbol=symbol,
                    qty=current_qty,
                    base_note="invalid_dividend_date_gaps",
                    events_observed=len(events),
                    last_dividend_date=unique_dates[-1].isoformat(),
                    median_dividend_amount=(
                        round(median_rate * current_qty, 2)
                        if median_rate is not None
                        else round(statistics.median(amounts), 2)
                    ),
                )
                symbols.append(row)
                total_30 += float(row.get("projected_next_30_days") or 0.0)
                total_60 += float(row.get("projected_next_60_days") or 0.0)
                total_90 += float(row.get("projected_next_90_days") or 0.0)
                continue
            cadence_info = self._cadence_from_median_gap(statistics.median(day_gaps))
            if cadence_info is None:
                median_rate = (
                    statistics.median(per_share_rates) if per_share_rates else None
                )
                row = _yield_based_projection(
                    symbol=symbol,
                    qty=current_qty,
                    base_note="unable_to_infer_cadence",
                    events_observed=len(events),
                    last_dividend_date=unique_dates[-1].isoformat(),
                    median_dividend_amount=(
                        round(median_rate * current_qty, 2)
                        if median_rate is not None
                        else round(statistics.median(amounts), 2)
                    ),
                )
                symbols.append(row)
                total_30 += float(row.get("projected_next_30_days") or 0.0)
                total_60 += float(row.get("projected_next_60_days") or 0.0)
                total_90 += float(row.get("projected_next_90_days") or 0.0)
                continue
            cadence_days, cadence_label = cadence_info
            median_per_share_rate = (
                statistics.median(per_share_rates) if per_share_rates else None
            )
            if median_per_share_rate is None or median_per_share_rate <= 0:
                # Last-resort fallback if per-share extraction is unavailable.
                median_per_share_rate = statistics.median(amounts) / current_qty
            median_dividend_amount = median_per_share_rate * current_qty
            last_dividend_date = unique_dates[-1]
            if median_dividend_amount <= 0:
                continue

            projected_dates, projected_30, projected_60, projected_90 = (
                _build_projected_dates_and_totals(
                    last_date=last_dividend_date,
                    cadence_days=cadence_days,
                    cash_amount=median_dividend_amount,
                )
            )

            if not projected_dates:
                continue

            symbols.append(
                {
                    "instrument": symbol,
                    "cadence_label": cadence_label,
                    "cadence_days": cadence_days,
                    "events_observed": len(events),
                    "last_dividend_date": last_dividend_date.isoformat(),
                    "median_dividend_amount": round(median_dividend_amount, 2),
                    "projected_next_30_days": round(projected_30, 2),
                    "projected_next_60_days": round(projected_60, 2),
                    "projected_next_90_days": round(projected_90, 2),
                    "projected_payout_dates": [d.isoformat() for d in projected_dates],
                    "forecast_note": None,
                    "yield_source": None,
                    "dividend_yield_used": None,
                    "per_share_distribution_rate_used": round(median_per_share_rate, 6),
                    "shares_used_for_projection": round(current_qty, 6),
                    "current_value": (
                        round(pos_current_value, 2) if pos_current_value is not None else None
                    ),
                    "estimated_cost_basis": (
                        round(pos_estimated_cost_basis, 2)
                        if pos_estimated_cost_basis is not None
                        else None
                    ),
                    "total_pl_pct_of_cost_basis": (
                        round(pos_total_pl_pct, 4) if pos_total_pl_pct is not None else None
                    ),
                }
            )
            total_30 += projected_30
            total_60 += projected_60
            total_90 += projected_90

        symbols.sort(
            key=lambda row: float(row.get("projected_next_90_days") or 0.0),
            reverse=True,
        )
        for row in symbols:
            shares = self._to_float(row.get("shares_used_for_projection"))
            current_value = self._to_float(row.get("current_value"))
            growth_rate = max(
                0.0,
                float(self._to_float(row.get("total_pl_pct_of_cost_basis")) or 0.0)
                / 100.0,
            )
            projected_30 = self._to_float(row.get("projected_next_30_days"))
            effective_30 = (
                (projected_30 / shares)
                if (
                    projected_30 is not None
                    and shares is not None
                    and shares > 0
                )
                else None
            )
            row["effective_30_day_distribution_rate"] = (
                round(effective_30, 6) if effective_30 is not None else None
            )
            row["projected_growth_dollars"] = (
                round(current_value * growth_rate, 2)
                if current_value is not None
                else None
            )

        total_current_value_for_optimization = sum(
            float(self._to_float(r.get("current_value")) or 0.0) for r in symbols
        )

        def _apply_optimization_pair(
            prefix: str,
            score_key: str,
            score_transform: str = "nonnegative",
        ) -> None:
            if total_current_value_for_optimization <= 0:
                for row in symbols:
                    row[f"{prefix}_target_shares"] = None
                    row[f"{prefix}_target_30_day_distribution"] = None
                    row[f"{prefix}_target_projected_growth_dollars"] = None
                return
            scored_rows: list[tuple[dict[str, Any], float, float, float]] = []
            for row in symbols:
                shares = self._to_float(row.get("shares_used_for_projection"))
                current_value = self._to_float(row.get("current_value"))
                effective_30 = self._to_float(
                    row.get("effective_30_day_distribution_rate")
                )
                price = (
                    (current_value / shares)
                    if (
                        shares is not None
                        and shares > 0
                        and current_value is not None
                        and current_value > 0
                    )
                    else None
                )
                raw_score = self._to_float(row.get(score_key))
                if raw_score is None:
                    score = 0.0
                elif score_transform == "nonnegative":
                    score = max(0.0, raw_score)
                else:
                    score = raw_score
                scored_rows.append((row, score, price or 0.0, effective_30 or 0.0))

            total_score = sum(s for _, s, _, _ in scored_rows)
            if total_score <= 0:
                # Fall back to current value weights when no positive score signal.
                total_score = sum(float(self._to_float(r.get("current_value")) or 0.0) for r in symbols)
                scored_rows = [
                    (
                        row,
                        float(self._to_float(row.get("current_value")) or 0.0),
                        price,
                        eff30,
                    )
                    for row, _, price, eff30 in scored_rows
                ]

            for row, score, price, effective_30 in scored_rows:
                growth_rate = max(
                    0.0,
                    float(self._to_float(row.get("total_pl_pct_of_cost_basis")) or 0.0)
                    / 100.0,
                )
                if price <= 0 or score <= 0 or total_score <= 0:
                    row[f"{prefix}_target_shares"] = 0.0
                    row[f"{prefix}_target_30_day_distribution"] = 0.0
                    row[f"{prefix}_target_projected_growth_dollars"] = 0.0
                    continue
                target_value = total_current_value_for_optimization * (score / total_score)
                target_shares = target_value / price
                target_30_day_distribution = target_shares * effective_30
                target_growth_dollars = target_value * growth_rate
                row[f"{prefix}_target_shares"] = round(target_shares, 6)
                row[f"{prefix}_target_30_day_distribution"] = round(
                    target_30_day_distribution, 2
                )
                row[f"{prefix}_target_projected_growth_dollars"] = round(
                    target_growth_dollars, 2
                )

        for row in symbols:
            income_score = float(self._to_float(row.get("effective_30_day_distribution_rate")) or 0.0)
            growth_score = float(self._to_float(row.get("total_pl_pct_of_cost_basis")) or 0.0)
            row["_income_score"] = max(0.0, income_score)
            row["_growth_score"] = max(0.0, growth_score)
            row["_balanced_score"] = (row["_income_score"] + row["_growth_score"]) / 2.0

        _apply_optimization_pair("income_optimized", "_income_score")
        _apply_optimization_pair("growth_optimized", "_growth_score")
        _apply_optimization_pair("balanced_optimized", "_balanced_score")

        for row in symbols:
            row.pop("_income_score", None)
            row.pop("_growth_score", None)
            row.pop("_balanced_score", None)

        return {
            "as_of_date": as_of_date.isoformat(),
            "totals": {
                "projected_next_30_days": round(total_30, 2),
                "projected_next_60_days": round(total_60, 2),
                "projected_next_90_days": round(total_90, 2),
            },
            "symbols": symbols,
        }

    def analyze(
        self,
        include_zero_positions: bool = False,
        forecast_as_of: date | None = None,
    ) -> PositionSummary:
        if not self.account_dir.is_dir():
            raise FileNotFoundError(f"Account directory not found: {self.account_dir}")

        per_symbol: dict[str, dict[str, Any]] = {}
        all_records: list[dict[str, Any]] = []

        for path in sorted(self.account_dir.glob("*.json")):
            # Skip prior outputs from this analyzer if present.
            if path.name.endswith(".positions.json"):
                continue

            rec = self._read_json(path)
            if not isinstance(rec, dict):
                continue
            all_records.append(rec)

        all_records.sort(key=self._record_sort_key)
        records_read = len(all_records)

        # Cash reconciliation views.
        by_code: dict[str, float] = {}
        dividend_events_by_symbol: dict[str, list[dict[str, Any]]] = {}
        for rec in all_records:
            code = str(rec.get("trans_code") or "").strip().upper() or "(BLANK)"
            amount = self._to_float(rec.get("amount"))
            if amount is None:
                continue
            by_code[code] = by_code.get(code, 0.0) + amount

        trade_cash_delta = by_code.get("BUY", 0.0) + by_code.get("SELL", 0.0)
        transfer_cash_delta = by_code.get("ACATI", 0.0) + by_code.get("ACATO", 0.0)
        transfer_adjusted_cash_delta = trade_cash_delta + transfer_cash_delta
        dividends_cash_delta = by_code.get("CDIV", 0.0) + by_code.get("MDIV", 0.0)
        options_cash_delta = by_code.get("STO", 0.0) + by_code.get("BTC", 0.0)
        fees_cash_delta = by_code.get("FEE", 0.0) + by_code.get("EMRF", 0.0)
        match_cash_delta = by_code.get("MTCH", 0.0)
        full_cash_delta = sum(by_code.values())
        cash_reconciliation = {
            "trade_cash_delta_buy_sell": round(trade_cash_delta, 2),
            "transfer_cash_delta_acati_acato": round(transfer_cash_delta, 2),
            "transfer_adjusted_cash_delta": round(transfer_adjusted_cash_delta, 2),
            "dividends_cash_delta": round(dividends_cash_delta, 2),
            "options_cash_delta": round(options_cash_delta, 2),
            "fees_cash_delta": round(fees_cash_delta, 2),
            "match_cash_delta": round(match_cash_delta, 2),
            "full_cash_delta_all_amount_rows": round(full_cash_delta, 2),
            "cash_delta_by_trans_code": {
                code: round(val, 2) for code, val in sorted(by_code.items())
            },
        }

        for rec in all_records:
            symbol = str(rec.get("instrument") or "").strip().upper()
            if not symbol:
                continue
            if symbol in self.excluded_symbols:
                continue

            trans_code = str(rec.get("trans_code") or "").strip().upper()
            qty = self._to_float(rec.get("quantity"))
            amount = self._to_float(rec.get("amount")) or 0.0

            if symbol not in per_symbol:
                per_symbol[symbol] = {
                    "quantity": 0.0,
                    "cost_basis": 0.0,
                    "buy_quantity": 0.0,
                    "sell_quantity": 0.0,
                    "cash_in": 0.0,
                    "cash_out": 0.0,
                    "net_cash_flow": 0.0,
                    "dividend_cash": 0.0,
                    "options_cash_net": 0.0,
                    "buy_sell_transactions": [],
                }

            p = per_symbol[symbol]
            p["net_cash_flow"] += amount

            if amount > 0:
                p["cash_in"] += amount
            elif amount < 0:
                p["cash_out"] += abs(amount)
            if trans_code in DIVIDEND_CODES:
                p["dividend_cash"] += amount
            if trans_code in OPTIONS_CODES:
                p["options_cash_net"] += amount
            if trans_code in DIVIDEND_CODES:
                dividend_date = self._parse_iso_date(rec.get("activity_date"))
                if dividend_date is not None and amount > 0:
                    div_price = self._to_float(rec.get("price"))
                    div_qty = qty if (qty is not None and qty > 0) else None
                    if div_qty is None:
                        div_qty_desc, div_price_desc = self._extract_div_qty_price(
                            rec.get("description")
                        )
                        if div_qty_desc is not None and div_qty_desc > 0:
                            div_qty = div_qty_desc
                        if div_price is None and div_price_desc is not None and div_price_desc > 0:
                            div_price = div_price_desc
                    if div_qty is None and float(p.get("quantity") or 0.0) > 0:
                        div_qty = float(p.get("quantity") or 0.0)
                    amount_per_share = (
                        div_price
                        if div_price is not None and div_price > 0
                        else ((amount / div_qty) if div_qty is not None and div_qty > 0 else None)
                    )
                    dividend_events_by_symbol.setdefault(symbol, []).append(
                        {
                            "date": dividend_date,
                            "amount_total": amount,
                            "shares_at_dividend": div_qty,
                            "amount_per_share": amount_per_share,
                        }
                    )
            if trans_code in {"BUY", "SELL"}:
                p["buy_sell_transactions"].append(
                    {
                        "activity_date": rec.get("activity_date"),
                        "process_date": rec.get("process_date"),
                        "settle_date": rec.get("settle_date"),
                        "trans_code": rec.get("trans_code"),
                        "quantity": rec.get("quantity"),
                        "price": rec.get("price"),
                        "amount": rec.get("amount"),
                        "description": rec.get("description"),
                    }
                )
            if qty is None:
                continue
            if qty <= 0:
                continue

            if trans_code in BUY_LIKE_CODES:
                p["buy_quantity"] += qty
                p["quantity"] += qty
                if amount < 0:
                    p["cost_basis"] += abs(amount)
            elif trans_code in SELL_LIKE_CODES:
                p["sell_quantity"] += qty
                if p["quantity"] > 0:
                    avg_cost = p["cost_basis"] / p["quantity"] if p["quantity"] else 0.0
                    reduce_qty = min(qty, p["quantity"])
                    p["cost_basis"] -= avg_cost * reduce_qty
                p["quantity"] = max(0.0, p["quantity"] - qty)

        positions: list[dict[str, Any]] = []
        for symbol in sorted(per_symbol.keys()):
            p = per_symbol[symbol]
            quantity_raw = p["quantity"]
            quantity = round(quantity_raw, 6)
            if (not include_zero_positions) and quantity <= 0:
                continue
            avg_cost = (p["cost_basis"] / quantity_raw) if quantity_raw > 0 else 0.0
            current_price = self._get_current_price(symbol)
            current_value = (quantity_raw * current_price) if current_price is not None else None
            unrealized_pl = (
                current_value - p["cost_basis"] if current_value is not None else None
            )
            all_in_pl = None
            if unrealized_pl is not None:
                all_in_pl = unrealized_pl + p["dividend_cash"] + p["options_cash_net"]
            cost_basis = max(0.0, p["cost_basis"])
            dividend_pct_of_cost_basis = (
                (p["dividend_cash"] / cost_basis) * 100 if cost_basis > 0 else None
            )
            total_pl_pct_of_cost_basis = (
                (all_in_pl / cost_basis) * 100
                if (all_in_pl is not None and cost_basis > 0)
                else None
            )
            positions.append(
                {
                    "instrument": symbol,
                    "quantity": quantity,
                    "estimated_cost_basis": round(cost_basis, 2),
                    "estimated_avg_cost": round(max(0.0, avg_cost), 4),
                    "current_price": round(current_price, 4) if current_price is not None else None,
                    "current_value": round(current_value, 2) if current_value is not None else None,
                    "unrealized_pl": round(unrealized_pl, 2) if unrealized_pl is not None else None,
                    "all_in_pl_including_dividends_and_options": (
                        round(all_in_pl, 2) if all_in_pl is not None else None
                    ),
                    "dividend_pct_of_cost_basis": (
                        round(dividend_pct_of_cost_basis, 4)
                        if dividend_pct_of_cost_basis is not None
                        else None
                    ),
                    "total_pl_pct_of_cost_basis": (
                        round(total_pl_pct_of_cost_basis, 4)
                        if total_pl_pct_of_cost_basis is not None
                        else None
                    ),
                    "buy_quantity": round(p["buy_quantity"], 6),
                    "sell_quantity": round(p["sell_quantity"], 6),
                    "cash_in": round(p["cash_in"], 2),
                    "cash_out": round(p["cash_out"], 2),
                    "net_cash_flow": round(p["net_cash_flow"], 2),
                    "dividend_sum": round(p["dividend_cash"], 2),
                    "dividend_cash": round(p["dividend_cash"], 2),
                    "options_cash_net": round(p["options_cash_net"], 2),
                    "buy_sell_transactions": p["buy_sell_transactions"],
                }
            )

        current_position_qty_by_symbol = {
            symbol: round(float(p.get("quantity") or 0.0), 6)
            for symbol, p in per_symbol.items()
            if round(float(p.get("quantity") or 0.0), 6) > 0
        }
        position_snapshot_by_symbol = {
            str(p.get("instrument") or "").strip().upper(): p for p in positions
        }

        cash_forecast = self._build_cash_forecast(
            dividend_events_by_symbol=dividend_events_by_symbol,
            current_position_qty_by_symbol=current_position_qty_by_symbol,
            position_snapshot_by_symbol=position_snapshot_by_symbol,
            as_of_date=forecast_as_of or date.today(),
        )

        return PositionSummary(
            account_name=self.account_name,
            records_read=records_read,
            positions=positions,
            cash_reconciliation=cash_reconciliation,
            cash_forecast=cash_forecast,
        )

    def _build_cash_monte_carlo_analysis(
        self,
        symbols: list[dict[str, Any]],
        simulations: int = 4000,
    ) -> dict[str, Any]:
        constraints = self.monte_carlo_constraints or {}
        base_total_value = sum(
            float(self._to_float(r.get("current_value")) or 0.0) for r in symbols
        )
        include_symbols_even_if_not_held = (
            constraints.get("include_symbols_even_if_not_held")
            if isinstance(constraints.get("include_symbols_even_if_not_held"), list)
            else []
        )
        existing_symbols = {
            str(r.get("instrument") or "").strip().upper() for r in symbols
        }
        symbols_for_universe = list(symbols)
        for sym in include_symbols_even_if_not_held:
            instrument = str(sym or "").strip().upper()
            if not instrument or instrument in existing_symbols:
                continue
            price = self._get_current_price(instrument)
            if price is None or price <= 0:
                continue
            annual_yield = self._get_stockanalysis_yield(instrument)
            if annual_yield is None or annual_yield <= 0:
                annual_yield = self._get_dividend_yield(instrument)
            projected_income_30 = (
                (price * annual_yield * (30.0 / 365.0))
                if annual_yield is not None and annual_yield > 0
                else 0.0
            )
            symbols_for_universe.append(
                {
                    "instrument": instrument,
                    "current_value": round(price, 6),
                    "projected_next_30_days": round(projected_income_30, 6),
                    "shares_used_for_projection": 1.0,
                    "effective_30_day_distribution_rate": round(projected_income_30, 8),
                    "_synthetic_monte_carlo_seed": True,
                }
            )
            existing_symbols.add(instrument)
        configured_simulations = self._parse_positive_int(
            constraints.get("simulation_count")
        )
        if configured_simulations is not None:
            simulations = configured_simulations
        sampling_alpha_scale = float(
            self._to_float(constraints.get("sampling_alpha_scale")) or 1.0
        )
        if sampling_alpha_scale <= 0:
            sampling_alpha_scale = 1.0
        growth_conf_weight = float(
            self._to_float(constraints.get("allocation_confidence_growth_weight")) or 0.45
        )
        dividend_conf_weight = float(
            self._to_float(constraints.get("allocation_confidence_dividend_weight")) or 0.55
        )
        conf_weight_sum = growth_conf_weight + dividend_conf_weight
        if conf_weight_sum <= 0:
            growth_conf_weight = 0.45
            dividend_conf_weight = 0.55
            conf_weight_sum = 1.0
        growth_conf_weight /= conf_weight_sum
        dividend_conf_weight /= conf_weight_sum
        conf_mult_min = float(
            self._to_float(constraints.get("allocation_confidence_multiplier_min")) or 0.8
        )
        conf_mult_max = float(
            self._to_float(constraints.get("allocation_confidence_multiplier_max")) or 1.2
        )
        if conf_mult_min <= 0 or conf_mult_max <= 0 or conf_mult_min > conf_mult_max:
            conf_mult_min = 0.8
            conf_mult_max = 1.2
        min_price_history_months = self._parse_nonnegative_float(
            constraints.get("min_price_history_months")
        )
        min_confidence_score = self._parse_pct(constraints.get("min_confidence_score"))
        low_confidence_score_threshold = self._parse_pct(
            constraints.get("low_confidence_score_threshold")
        )
        max_low_conf_total_weight_frac = (
            float(
                self._to_float(constraints.get("max_low_confidence_total_weight_pct"))
                or 0.0
            )
            / 100.0
        )
        max_low_conf_total_weight_frac = (
            max(0.0, min(1.0, max_low_conf_total_weight_frac))
            if max_low_conf_total_weight_frac > 0
            else 0.0
        )
        max_annualized_growth_pct_for_projection = self._parse_pct(
            constraints.get("max_annualized_growth_pct_for_projection")
        )
        max_annualized_growth_rate_for_projection = (
            (max_annualized_growth_pct_for_projection / 100.0)
            if max_annualized_growth_pct_for_projection is not None
            else None
        )

        universe: list[dict[str, Any]] = []
        filtered_out_price_history_count = 0
        filtered_out_confidence_count = 0
        for row in symbols_for_universe:
            instrument = str(row.get("instrument") or "").strip().upper()
            current_value = self._to_float(row.get("current_value"))
            projected_income_30 = self._to_float(row.get("projected_next_30_days"))
            current_shares = self._to_float(row.get("shares_used_for_projection"))
            effective_30_rate = self._to_float(
                row.get("effective_30_day_distribution_rate")
            )
            if (
                (not instrument)
                or
                current_value is None
                or projected_income_30 is None
                or current_value <= 0
            ):
                continue
            price = (
                (current_value / current_shares)
                if (
                    current_shares is not None
                    and current_shares > 0
                    and current_value > 0
                )
                else None
            )
            try:
                growth_result = PriceHistoryGrowthAnalyzer(instrument).analyze()
            except Exception:
                growth_result = None
            growth_confidence_score = max(
                0.0,
                min(
                    100.0,
                    float(
                        self._to_float(
                            growth_result.confidence_score
                            if growth_result is not None
                            else 0.0
                        )
                        or 0.0
                    ),
                ),
            )
            annualized_growth_rate = max(
                0.0,
                float(
                    self._to_float(
                        growth_result.annualized_growth_pct
                        if growth_result is not None
                        else 0.0
                    )
                    or 0.0
                )
                / 100.0,
            )
            if max_annualized_growth_rate_for_projection is not None:
                annualized_growth_rate = min(
                    annualized_growth_rate,
                    max_annualized_growth_rate_for_projection,
                )
            # Convert annualized growth to an expected 30-day growth-per-dollar figure.
            # Confidence affects allocation preference, not the expected return itself.
            growth_per_dollar_30 = annualized_growth_rate * (30.0 / 365.0)
            try:
                div_signal = DistributionHistoryComparison(instrument).dividend_history_signal()
            except Exception:
                div_signal = {
                    "source": "none",
                    "rate_30_day_per_share": 0.0,
                    "confidence_score": 0.0,
                    "history_months": 0.0,
                }
            div_rate_30_per_share = max(
                0.0,
                float(
                    self._to_float(div_signal.get("rate_30_day_per_share")) or 0.0
                ),
            )
            div_conf = max(
                0.0,
                min(
                    100.0,
                    float(self._to_float(div_signal.get("confidence_score")) or 0.0),
                ),
            )
            # Keep expected dividend signal unscaled by confidence; confidence will
            # influence allocation percentages instead of suppressing expected values.
            if price is not None and price > 0 and div_rate_30_per_share > 0:
                income_per_dollar = div_rate_30_per_share / price
            else:
                income_per_dollar = projected_income_30 / current_value
            has_div_signal = bool(price is not None and price > 0 and div_rate_30_per_share > 0)
            combined_allocation_confidence = (
                (growth_conf_weight * growth_confidence_score)
                + (dividend_conf_weight * div_conf)
                if has_div_signal
                else growth_confidence_score
            )
            price_history_months = (
                round(
                    float(
                        self._to_float(
                            growth_result.years_of_history if growth_result is not None else 0.0
                        )
                        or 0.0
                    )
                    * 12.0,
                    2,
                )
            )
            if (
                min_price_history_months is not None
                and price_history_months < float(min_price_history_months)
            ):
                filtered_out_price_history_count += 1
                continue
            if (
                min_confidence_score is not None
                and combined_allocation_confidence < float(min_confidence_score)
            ):
                filtered_out_confidence_count += 1
                continue
            # Mildly bias allocations toward higher-confidence symbols without overpowering
            # the underlying expected income/growth signals.
            allocation_confidence_multiplier = conf_mult_min + (
                (conf_mult_max - conf_mult_min)
                * (combined_allocation_confidence / 100.0)
            )
            universe.append(
                {
                    "instrument": instrument,
                    "price": price,
                    "effective_30_day_distribution_rate": effective_30_rate,
                    "income_per_dollar_30": income_per_dollar,
                    "growth_per_dollar": growth_per_dollar_30,
                    "price_history_growth_pct": (
                        growth_result.growth_pct if growth_result is not None else 0.0
                    ),
                    "price_history_annualized_growth_pct": (
                        growth_result.annualized_growth_pct
                        if growth_result is not None
                        else 0.0
                    ),
                    "price_history_months": price_history_months,
                    "price_history_confidence_score": (
                        growth_result.confidence_score
                        if growth_result is not None
                        else 0.0
                    ),
                    "derived_price_growth_pct": round(
                        annualized_growth_rate * 100.0, 4
                    ),
                    "dividend_history_rate_30_day_per_share": round(
                        div_rate_30_per_share, 8
                    ),
                    "dividend_history_confidence_score": round(div_conf, 2),
                    "dividend_history_months": round(
                        float(self._to_float(div_signal.get("history_months")) or 0.0), 2
                    ),
                    "dividend_history_source": str(div_signal.get("source") or "none"),
                    "allocation_confidence_score": round(combined_allocation_confidence, 2),
                    "allocation_confidence_multiplier": round(
                        allocation_confidence_multiplier, 6
                    ),
                }
            )
        total_value = base_total_value
        if len(universe) < 2 or total_value <= 0:
            return {
                "simulations": 0,
                "total_value": round(total_value, 2),
                "symbol_count": len(universe),
                "top_scenarios": [],
                "frontier_points": [],
            }

        max_symbol_weight_frac = (
            float(self._to_float(constraints.get("max_symbol_weight_pct")) or 0.0) / 100.0
        )
        max_symbol_weight_frac = (
            max(0.0, min(1.0, max_symbol_weight_frac))
            if max_symbol_weight_frac > 0
            else 1.0
        )
        confidence_tier_caps_raw = (
            constraints.get("confidence_tier_symbol_caps")
            if isinstance(constraints.get("confidence_tier_symbol_caps"), list)
            else []
        )
        confidence_tier_caps: list[tuple[float, float]] = []
        for row in confidence_tier_caps_raw:
            if not isinstance(row, dict):
                continue
            max_conf = self._to_float(row.get("max_confidence_score"))
            max_weight_pct = self._to_float(row.get("max_weight_pct"))
            if (
                max_conf is None
                or max_weight_pct is None
                or max_conf < 0
                or max_conf > 100
                or max_weight_pct <= 0
            ):
                continue
            confidence_tier_caps.append(
                (float(max_conf), max(0.0, min(1.0, float(max_weight_pct) / 100.0)))
            )
        confidence_tier_caps.sort(key=lambda x: x[0])
        high_conf_cap_row = (
            constraints.get("high_confidence_symbol_cap")
            if isinstance(constraints.get("high_confidence_symbol_cap"), dict)
            else None
        )
        high_conf_min_score = (
            self._to_float(high_conf_cap_row.get("min_confidence_score"))
            if high_conf_cap_row is not None
            else None
        )
        high_conf_cap_frac = (
            (
                float(self._to_float(high_conf_cap_row.get("max_weight_pct")) or 0.0)
                / 100.0
            )
            if high_conf_cap_row is not None
            else None
        )
        if high_conf_cap_frac is not None:
            high_conf_cap_frac = max(0.0, min(1.0, high_conf_cap_frac))
        symbol_caps: list[float] = []
        tier_band_specs: list[dict[str, Any]] = []
        prev_bound: float | None = None
        for tier_max_conf, tier_cap in confidence_tier_caps:
            tier_band_specs.append(
                {
                    "min_conf_exclusive": prev_bound,
                    "max_conf_inclusive": tier_max_conf,
                    "cap_frac": tier_cap,
                    "symbol_indices": [],
                }
            )
            prev_bound = tier_max_conf
        for i, sym in enumerate(universe):
            sym_cap = max_symbol_weight_frac
            conf_score = float(
                self._to_float(sym.get("allocation_confidence_score")) or 0.0
            )
            matched_tier_idx: int | None = None
            for tier_idx, (tier_max_conf, tier_cap) in enumerate(confidence_tier_caps):
                if conf_score <= tier_max_conf:
                    sym_cap = min(sym_cap, tier_cap)
                    matched_tier_idx = tier_idx
                    break
            if (
                high_conf_min_score is not None
                and high_conf_cap_frac is not None
                and conf_score >= float(high_conf_min_score)
            ):
                sym_cap = max(sym_cap, high_conf_cap_frac)
            symbol_caps.append(sym_cap)
            if matched_tier_idx is not None and matched_tier_idx < len(tier_band_specs):
                tier_band_specs[matched_tier_idx]["symbol_indices"].append(i)
        theme_caps_indexed: list[tuple[set[int], float]] = []
        for cap in constraints.get("theme_caps") or []:
            if not isinstance(cap, dict):
                continue
            cap_frac = float(self._to_float(cap.get("max_weight_pct")) or 0.0) / 100.0
            if cap_frac <= 0:
                continue
            symbols_in_cap = {
                str(s or "").strip().upper() for s in (cap.get("symbols") or [])
            }
            if not symbols_in_cap:
                continue
            idx = {
                i
                for i, sym in enumerate(universe)
                if str(sym.get("instrument") or "").strip().upper() in symbols_in_cap
            }
            if idx:
                theme_caps_indexed.append((idx, max(0.0, min(1.0, cap_frac))))
        max_income_contrib_frac = (
            float(self._to_float(constraints.get("max_income_contribution_pct")) or 0.0)
            / 100.0
        )
        max_growth_contrib_frac = (
            float(self._to_float(constraints.get("max_growth_contribution_pct")) or 0.0)
            / 100.0
        )
        max_top_5_weight_frac = (
            float(self._to_float(constraints.get("max_top_5_weight_pct")) or 0.0) / 100.0
        )
        max_top_10_weight_frac = (
            float(self._to_float(constraints.get("max_top_10_weight_pct")) or 0.0) / 100.0
        )
        low_conf_idx = {
            i
            for i, sym in enumerate(universe)
            if (
                low_confidence_score_threshold is not None
                and float(self._to_float(sym.get("allocation_confidence_score")) or 0.0)
                < float(low_confidence_score_threshold)
            )
        }
        scenario_constraints = (
            constraints.get("scenario_constraints")
            if isinstance(constraints.get("scenario_constraints"), dict)
            else {}
        )
        shock_pcts = (
            constraints.get("top_concentration_shocks_pct")
            if isinstance(constraints.get("top_concentration_shocks_pct"), list)
            else [20.0, 30.0]
        )
        shock_pcts = [
            float(self._to_float(v) or 0.0)
            for v in shock_pcts
            if float(self._to_float(v) or 0.0) > 0.0
        ]
        if not shock_pcts:
            shock_pcts = [20.0, 30.0]

        rng = random.Random(42)
        runs: list[dict[str, Any]] = []
        attempted_runs = 0
        symbol_cap_hit_count = 0
        theme_cap_hit_count = 0
        rejected_income_contrib_count = 0
        rejected_growth_contrib_count = 0
        rejected_low_conf_weight_count = 0
        rejected_top_5_weight_count = 0
        rejected_top_10_weight_count = 0
        tier_hit_stats_internal: list[dict[str, Any]] = [
            {
                "min_conf_exclusive": t.get("min_conf_exclusive"),
                "max_conf_inclusive": t.get("max_conf_inclusive"),
                "cap_frac": t.get("cap_frac"),
                "symbol_count": len(t.get("symbol_indices") or []),
                "hit_run_count": 0,
                "hit_symbol_occurrences": 0,
                "cumulative_excess_weight_pct": 0.0,
            }
            for t in tier_band_specs
        ]
        for i in range(simulations):
            attempted_runs += 1
            draws = [
                rng.gammavariate(
                    max(
                        0.1,
                        float(
                            self._to_float(
                                universe[j].get("allocation_confidence_multiplier")
                            )
                            or 1.0
                        )
                        * sampling_alpha_scale,
                    ),
                    1.0,
                )
                for j in range(len(universe))
            ]
            raw_weights = self._normalize_weights(draws)
            if max_symbol_weight_frac < 1.0 and any(
                w > (max_symbol_weight_frac + 1e-12) for w in raw_weights
            ):
                symbol_cap_hit_count += 1
            for tier_idx, tier in enumerate(tier_band_specs):
                idx_list = tier.get("symbol_indices") or []
                if not idx_list:
                    continue
                cap_frac = float(self._to_float(tier.get("cap_frac")) or 0.0)
                if cap_frac <= 0:
                    continue
                over = [
                    (raw_weights[idx] - cap_frac)
                    for idx in idx_list
                    if raw_weights[idx] > (cap_frac + 1e-12)
                ]
                if not over:
                    continue
                stat = tier_hit_stats_internal[tier_idx]
                stat["hit_run_count"] = int(stat.get("hit_run_count") or 0) + 1
                stat["hit_symbol_occurrences"] = int(
                    stat.get("hit_symbol_occurrences") or 0
                ) + len(over)
                stat["cumulative_excess_weight_pct"] = float(
                    self._to_float(stat.get("cumulative_excess_weight_pct")) or 0.0
                ) + (sum(over) * 100.0)
            if theme_caps_indexed:
                raw_theme_hit = False
                for member_idx, cap in theme_caps_indexed:
                    group_weight = sum(raw_weights[idx] for idx in member_idx)
                    if group_weight > (cap + 1e-12):
                        raw_theme_hit = True
                        break
                if raw_theme_hit:
                    theme_cap_hit_count += 1
            weights = self._apply_symbol_weight_caps(raw_weights, symbol_caps)
            weights = self._apply_theme_caps(weights, theme_caps_indexed)
            if max_top_5_weight_frac > 0:
                top_5_weight = sum(sorted(weights, reverse=True)[:5])
                if top_5_weight > max_top_5_weight_frac:
                    rejected_top_5_weight_count += 1
                    continue
            if max_top_10_weight_frac > 0:
                top_10_weight = sum(sorted(weights, reverse=True)[:10])
                if top_10_weight > max_top_10_weight_frac:
                    rejected_top_10_weight_count += 1
                    continue
            expected_income_30 = total_value * sum(
                weights[j] * universe[j]["income_per_dollar_30"] for j in range(len(universe))
            )
            expected_growth = total_value * sum(
                weights[j] * universe[j]["growth_per_dollar"] for j in range(len(universe))
            )
            low_confidence_weight = (
                sum(weights[j] for j in low_conf_idx) if low_conf_idx else 0.0
            )
            if (
                max_low_conf_total_weight_frac > 0
                and low_confidence_weight > max_low_conf_total_weight_frac
            ):
                rejected_low_conf_weight_count += 1
                continue
            if expected_income_30 > 0 and max_income_contrib_frac > 0:
                max_income_contrib = max(
                    (
                        (total_value * weights[j] * universe[j]["income_per_dollar_30"])
                        / expected_income_30
                    )
                    for j in range(len(universe))
                )
                if max_income_contrib > max_income_contrib_frac:
                    rejected_income_contrib_count += 1
                    continue
            if expected_growth > 0 and max_growth_contrib_frac > 0:
                max_growth_contrib = max(
                    (
                        (total_value * weights[j] * universe[j]["growth_per_dollar"])
                        / expected_growth
                    )
                    for j in range(len(universe))
                )
                if max_growth_contrib > max_growth_contrib_frac:
                    rejected_growth_contrib_count += 1
                    continue
            runs.append(
                {
                    "run_id": i + 1,
                    "weights": weights,
                    "expected_income_30": expected_income_30,
                    "expected_growth": expected_growth,
                    "low_confidence_weight": low_confidence_weight,
                }
            )
        if not runs:
            return {
                "simulations": 0,
                "attempted_simulations": attempted_runs,
                "total_value": round(total_value, 2),
                "symbol_count": len(universe),
                "constraints_applied": {
                    "simulation_count": simulations,
                    "sampling_alpha_scale": sampling_alpha_scale,
                    "include_symbols_even_if_not_held": include_symbols_even_if_not_held,
                    "max_symbol_weight_pct": constraints.get("max_symbol_weight_pct"),
                    "max_top_5_weight_pct": constraints.get("max_top_5_weight_pct"),
                    "max_top_10_weight_pct": constraints.get("max_top_10_weight_pct"),
                    "confidence_tier_symbol_caps": constraints.get(
                        "confidence_tier_symbol_caps"
                    )
                    or [],
                    "high_confidence_symbol_cap": constraints.get(
                        "high_confidence_symbol_cap"
                    ),
                    "max_income_contribution_pct": constraints.get(
                        "max_income_contribution_pct"
                    ),
                    "max_growth_contribution_pct": constraints.get(
                        "max_growth_contribution_pct"
                    ),
                    "allocation_confidence_growth_weight": round(growth_conf_weight, 4),
                    "allocation_confidence_dividend_weight": round(
                        dividend_conf_weight, 4
                    ),
                    "allocation_confidence_multiplier_min": conf_mult_min,
                    "allocation_confidence_multiplier_max": conf_mult_max,
                    "min_accepted_run_pct_target": constraints.get(
                        "min_accepted_run_pct_target"
                    ),
                    "min_price_history_months": constraints.get("min_price_history_months"),
                    "min_confidence_score": constraints.get("min_confidence_score"),
                    "low_confidence_score_threshold": constraints.get(
                        "low_confidence_score_threshold"
                    ),
                    "max_low_confidence_total_weight_pct": constraints.get(
                        "max_low_confidence_total_weight_pct"
                    ),
                    "max_annualized_growth_pct_for_projection": constraints.get(
                        "max_annualized_growth_pct_for_projection"
                    ),
                    "top_concentration_shocks_pct": shock_pcts,
                    "scenario_constraints": scenario_constraints,
                    "theme_caps": constraints.get("theme_caps") or [],
                },
                "constraint_hit_stats": {
                    "symbol_cap_hit_count": symbol_cap_hit_count,
                    "symbol_cap_hit_pct_of_attempts": round(
                        (symbol_cap_hit_count / attempted_runs * 100.0)
                        if attempted_runs > 0
                        else 0.0,
                        2,
                    ),
                    "theme_cap_hit_count": theme_cap_hit_count,
                    "theme_cap_hit_pct_of_attempts": round(
                        (theme_cap_hit_count / attempted_runs * 100.0)
                        if attempted_runs > 0
                        else 0.0,
                        2,
                    ),
                    "rejected_growth_contribution_count": rejected_growth_contrib_count,
                    "rejected_growth_contribution_pct_of_attempts": round(
                        (rejected_growth_contrib_count / attempted_runs * 100.0)
                        if attempted_runs > 0
                        else 0.0,
                        2,
                    ),
                    "rejected_income_contribution_count": rejected_income_contrib_count,
                    "rejected_income_contribution_pct_of_attempts": round(
                        (rejected_income_contrib_count / attempted_runs * 100.0)
                        if attempted_runs > 0
                        else 0.0,
                        2,
                    ),
                    "rejected_low_confidence_weight_count": rejected_low_conf_weight_count,
                    "rejected_low_confidence_weight_pct_of_attempts": round(
                        (rejected_low_conf_weight_count / attempted_runs * 100.0)
                        if attempted_runs > 0
                        else 0.0,
                        2,
                    ),
                    "rejected_top_5_weight_count": rejected_top_5_weight_count,
                    "rejected_top_5_weight_pct_of_attempts": round(
                        (rejected_top_5_weight_count / attempted_runs * 100.0)
                        if attempted_runs > 0
                        else 0.0,
                        2,
                    ),
                    "rejected_top_10_weight_count": rejected_top_10_weight_count,
                    "rejected_top_10_weight_pct_of_attempts": round(
                        (rejected_top_10_weight_count / attempted_runs * 100.0)
                        if attempted_runs > 0
                        else 0.0,
                        2,
                    ),
                    "filtered_out_min_price_history_count": filtered_out_price_history_count,
                    "filtered_out_min_confidence_count": filtered_out_confidence_count,
                    "accepted_run_count": 0,
                    "accepted_run_pct_of_attempts": 0.0,
                    "accepted_run_pct_target": constraints.get("min_accepted_run_pct_target"),
                    "accepted_run_pct_below_target": (
                        (
                            0.0
                            < float(
                                self._to_float(
                                    constraints.get("min_accepted_run_pct_target")
                                )
                                or 0.0
                            )
                        )
                        if self._to_float(constraints.get("min_accepted_run_pct_target"))
                        is not None
                        else False
                    ),
                    "confidence_tier_hit_stats": [
                        {
                            "min_conf_exclusive": row.get("min_conf_exclusive"),
                            "max_conf_inclusive": row.get("max_conf_inclusive"),
                            "cap_pct": round(
                                (float(self._to_float(row.get("cap_frac")) or 0.0) * 100.0),
                                4,
                            ),
                            "symbol_count": int(row.get("symbol_count") or 0),
                            "hit_run_count": int(row.get("hit_run_count") or 0),
                            "hit_run_pct_of_attempts": round(
                                (
                                    (int(row.get("hit_run_count") or 0) / attempted_runs)
                                    * 100.0
                                    if attempted_runs > 0
                                    else 0.0
                                ),
                                2,
                            ),
                            "hit_symbol_occurrences": int(
                                row.get("hit_symbol_occurrences") or 0
                            ),
                            "avg_excess_weight_pct_per_hit_run": round(
                                (
                                    float(
                                        self._to_float(
                                            row.get("cumulative_excess_weight_pct")
                                        )
                                        or 0.0
                                    )
                                    / int(row.get("hit_run_count") or 1)
                                )
                                if int(row.get("hit_run_count") or 0) > 0
                                else 0.0,
                                4,
                            ),
                        }
                        for row in tier_hit_stats_internal
                    ],
                },
                "top_scenarios": [],
                "frontier_points": [],
            }

        incomes = [r["expected_income_30"] for r in runs]
        growths = [r["expected_growth"] for r in runs]
        income_mean = statistics.mean(incomes)
        growth_mean = statistics.mean(growths)
        income_stdev = statistics.stdev(incomes) if len(incomes) > 1 else 1.0
        growth_stdev = statistics.stdev(growths) if len(growths) > 1 else 1.0
        income_stdev = income_stdev if income_stdev > 0 else 1.0
        growth_stdev = growth_stdev if growth_stdev > 0 else 1.0

        for r in runs:
            z_income = (r["expected_income_30"] - income_mean) / income_stdev
            z_growth = (r["expected_growth"] - growth_mean) / growth_stdev
            r["balanced_score"] = z_income + z_growth

        def _allocation_str(weights: list[float], top_n: int = 6) -> str:
            ranked = sorted(
                ((universe[i]["instrument"], w) for i, w in enumerate(weights)),
                key=lambda x: x[1],
                reverse=True,
            )[:top_n]
            return ", ".join(f"{sym}:{round(w * 100, 1)}%" for sym, w in ranked)

        def _allocation_rows(weights: list[float]) -> list[dict[str, Any]]:
            rows: list[dict[str, Any]] = []
            for i, w in enumerate(weights):
                if w <= 0:
                    continue
                sym = universe[i]
                target_value = total_value * w
                expected_income_30 = target_value * sym["income_per_dollar_30"]
                expected_growth = target_value * sym["growth_per_dollar"]
                target_shares = (
                    target_value / sym["price"]
                    if sym.get("price") is not None and sym["price"] > 0
                    else None
                )
                rows.append(
                    {
                        "instrument": sym["instrument"],
                        "weight_pct": round(w * 100.0, 4),
                        "target_value": round(target_value, 2),
                        "target_shares": (
                            round(target_shares, 6)
                            if target_shares is not None
                            else None
                        ),
                        "expected_30_day_distribution": round(expected_income_30, 2),
                        "expected_growth_dollars": round(expected_growth, 2),
                        "price_history_months": sym.get("price_history_months"),
                        "derived_price_growth_pct": sym.get("derived_price_growth_pct"),
                        "price_history_confidence_score": sym.get(
                            "price_history_confidence_score"
                        ),
                        "dividend_history_rate_30_day_per_share": sym.get(
                            "dividend_history_rate_30_day_per_share"
                        ),
                        "dividend_history_confidence_score": sym.get(
                            "dividend_history_confidence_score"
                        ),
                        "dividend_history_months": sym.get("dividend_history_months"),
                        "dividend_history_source": sym.get("dividend_history_source"),
                        "allocation_confidence_score": sym.get(
                            "allocation_confidence_score"
                        ),
                        "allocation_confidence_multiplier": sym.get(
                            "allocation_confidence_multiplier"
                        ),
                    }
                )
            rows.sort(
                key=lambda r: float(r.get("expected_30_day_distribution") or 0.0),
                reverse=True,
            )
            return rows

        def _top_concentration_stress(weights: list[float], top_n: int = 5) -> dict[str, Any]:
            ranked_idx = sorted(
                range(len(weights)),
                key=lambda i: weights[i],
                reverse=True,
            )[:top_n]
            shocked_symbols = [universe[i]["instrument"] for i in ranked_idx]
            shocked_weight = sum(weights[i] for i in ranked_idx)
            tests: list[dict[str, Any]] = []
            for shock_pct in shock_pcts:
                shock_frac = shock_pct / 100.0
                drawdown_dollars = total_value * sum(weights[i] * shock_frac for i in ranked_idx)
                tests.append(
                    {
                        "shock_pct": round(shock_pct, 4),
                        "drawdown_dollars": round(drawdown_dollars, 2),
                        "drawdown_pct_of_portfolio": round(
                            ((drawdown_dollars / total_value) * 100.0) if total_value > 0 else 0.0,
                            4,
                        ),
                    }
                )
            return {
                "top_n": top_n,
                "symbols": shocked_symbols,
                "combined_weight_pct": round(shocked_weight * 100.0, 4),
                "tests": tests,
            }

        def _scenario_filter(run: dict[str, Any], key: str) -> bool:
            row = scenario_constraints.get(key)
            if not isinstance(row, dict):
                return True
            min_income = self._parse_nonnegative_float(row.get("min_income_30_day_distribution"))
            min_growth = self._parse_nonnegative_float(row.get("min_growth_30_day_dollars"))
            max_low_conf = self._parse_pct(row.get("max_low_confidence_weight_pct"))
            if (
                min_income is not None
                and float(self._to_float(run.get("expected_income_30")) or 0.0) < min_income
            ):
                return False
            if (
                min_growth is not None
                and float(self._to_float(run.get("expected_growth")) or 0.0) < min_growth
            ):
                return False
            if max_low_conf is not None:
                low_conf_weight_pct = (
                    float(self._to_float(run.get("low_confidence_weight")) or 0.0) * 100.0
                )
                if low_conf_weight_pct > max_low_conf:
                    return False
            return True

        def _pick_best(
            runs_in: list[dict[str, Any]],
            metric: str,
            scenario_key: str,
        ) -> tuple[dict[str, Any], bool]:
            eligible = [r for r in runs_in if _scenario_filter(r, scenario_key)]
            if eligible:
                return max(eligible, key=lambda r: r[metric]), False
            return max(runs_in, key=lambda r: r[metric]), True

        max_income, max_income_fallback = _pick_best(
            runs, "expected_income_30", "max_income"
        )
        max_growth, max_growth_fallback = _pick_best(
            runs, "expected_growth", "max_growth"
        )
        balanced, balanced_fallback = _pick_best(runs, "balanced_score", "balanced")

        sorted_by_income = sorted(runs, key=lambda r: r["expected_income_30"])
        q_idx = max(0, int(0.75 * (len(sorted_by_income) - 1)))
        income_floor = sorted_by_income[q_idx]["expected_income_30"]
        growth_floor = sorted(runs, key=lambda r: r["expected_growth"])[q_idx]["expected_growth"]
        high_income_candidates = [r for r in runs if r["expected_income_30"] >= income_floor]
        high_growth_candidates = [r for r in runs if r["expected_growth"] >= growth_floor]
        growth_given_income_floor, growth_given_income_floor_fallback = _pick_best(
            high_income_candidates,
            "expected_growth",
            "growth_given_income_floor",
        )
        income_given_growth_floor, income_given_growth_floor_fallback = _pick_best(
            high_growth_candidates,
            "expected_income_30",
            "income_given_growth_floor",
        )

        scenario_rows = [
            ("max_income", max_income, max_income_fallback),
            ("max_growth", max_growth, max_growth_fallback),
            ("balanced", balanced, balanced_fallback),
            (
                "growth_given_income_floor",
                growth_given_income_floor,
                growth_given_income_floor_fallback,
            ),
            (
                "income_given_growth_floor",
                income_given_growth_floor,
                income_given_growth_floor_fallback,
            ),
        ]

        top_scenarios = [
            {
                "scenario": name,
                "expected_30_day_distribution": round(run["expected_income_30"], 2),
                "expected_growth_dollars": round(run["expected_growth"], 2),
                "allocation_top_weights": _allocation_str(run["weights"]),
                "top_5_weight_pct": round(sum(sorted(run["weights"], reverse=True)[:5]) * 100.0, 4),
                "top_10_weight_pct": round(
                    sum(sorted(run["weights"], reverse=True)[:10]) * 100.0, 4
                ),
                "allocation_rows": _allocation_rows(run["weights"]),
                "scenario_constraints_applied": scenario_constraints.get(name) or {},
                "scenario_constraint_unmet_fallback": bool(fallback),
                "top_concentration_stress": _top_concentration_stress(run["weights"]),
            }
            for name, run, fallback in scenario_rows
        ]

        # Efficient frontier proxy: for each income decile, keep max growth run.
        frontier_points: list[dict[str, Any]] = []
        n = len(sorted_by_income)
        for decile in range(1, 11):
            start = int((decile - 1) * n / 10)
            end = int(decile * n / 10)
            bucket = sorted_by_income[start:max(end, start + 1)]
            if not bucket:
                continue
            best = max(bucket, key=lambda r: r["expected_growth"])
            frontier_points.append(
                {
                    "income_decile": decile,
                    "expected_30_day_distribution": round(best["expected_income_30"], 2),
                    "expected_growth_dollars": round(best["expected_growth"], 2),
                    "allocation_top_weights": _allocation_str(best["weights"], top_n=4),
                    "allocation_rows": _allocation_rows(best["weights"]),
                }
            )

        return {
            "simulations": len(runs),
            "attempted_simulations": attempted_runs,
            "total_value": round(total_value, 2),
            "symbol_count": len(universe),
            "constraints_applied": {
                "simulation_count": simulations,
                "sampling_alpha_scale": sampling_alpha_scale,
                "include_symbols_even_if_not_held": include_symbols_even_if_not_held,
                "max_symbol_weight_pct": constraints.get("max_symbol_weight_pct"),
                "max_top_5_weight_pct": constraints.get("max_top_5_weight_pct"),
                "max_top_10_weight_pct": constraints.get("max_top_10_weight_pct"),
                "confidence_tier_symbol_caps": constraints.get(
                    "confidence_tier_symbol_caps"
                )
                or [],
                "high_confidence_symbol_cap": constraints.get("high_confidence_symbol_cap"),
                "max_income_contribution_pct": constraints.get(
                    "max_income_contribution_pct"
                ),
                "max_growth_contribution_pct": constraints.get(
                    "max_growth_contribution_pct"
                ),
                "allocation_confidence_growth_weight": round(growth_conf_weight, 4),
                "allocation_confidence_dividend_weight": round(dividend_conf_weight, 4),
                "allocation_confidence_multiplier_min": conf_mult_min,
                "allocation_confidence_multiplier_max": conf_mult_max,
                "min_accepted_run_pct_target": constraints.get(
                    "min_accepted_run_pct_target"
                ),
                "min_price_history_months": constraints.get("min_price_history_months"),
                "min_confidence_score": constraints.get("min_confidence_score"),
                "low_confidence_score_threshold": constraints.get(
                    "low_confidence_score_threshold"
                ),
                "max_low_confidence_total_weight_pct": constraints.get(
                    "max_low_confidence_total_weight_pct"
                ),
                "max_annualized_growth_pct_for_projection": constraints.get(
                    "max_annualized_growth_pct_for_projection"
                ),
                "top_concentration_shocks_pct": shock_pcts,
                "scenario_constraints": scenario_constraints,
                "theme_caps": constraints.get("theme_caps") or [],
            },
            "constraint_hit_stats": {
                "symbol_cap_hit_count": symbol_cap_hit_count,
                "symbol_cap_hit_pct_of_attempts": round(
                    (symbol_cap_hit_count / attempted_runs * 100.0)
                    if attempted_runs > 0
                    else 0.0,
                    2,
                ),
                "theme_cap_hit_count": theme_cap_hit_count,
                "theme_cap_hit_pct_of_attempts": round(
                    (theme_cap_hit_count / attempted_runs * 100.0)
                    if attempted_runs > 0
                    else 0.0,
                    2,
                ),
                "rejected_growth_contribution_count": rejected_growth_contrib_count,
                "rejected_growth_contribution_pct_of_attempts": round(
                    (rejected_growth_contrib_count / attempted_runs * 100.0)
                    if attempted_runs > 0
                    else 0.0,
                    2,
                ),
                "rejected_income_contribution_count": rejected_income_contrib_count,
                "rejected_income_contribution_pct_of_attempts": round(
                    (rejected_income_contrib_count / attempted_runs * 100.0)
                    if attempted_runs > 0
                    else 0.0,
                    2,
                ),
                "rejected_low_confidence_weight_count": rejected_low_conf_weight_count,
                "rejected_low_confidence_weight_pct_of_attempts": round(
                    (rejected_low_conf_weight_count / attempted_runs * 100.0)
                    if attempted_runs > 0
                    else 0.0,
                    2,
                ),
                "rejected_top_5_weight_count": rejected_top_5_weight_count,
                "rejected_top_5_weight_pct_of_attempts": round(
                    (rejected_top_5_weight_count / attempted_runs * 100.0)
                    if attempted_runs > 0
                    else 0.0,
                    2,
                ),
                "rejected_top_10_weight_count": rejected_top_10_weight_count,
                "rejected_top_10_weight_pct_of_attempts": round(
                    (rejected_top_10_weight_count / attempted_runs * 100.0)
                    if attempted_runs > 0
                    else 0.0,
                    2,
                ),
                "filtered_out_min_price_history_count": filtered_out_price_history_count,
                "filtered_out_min_confidence_count": filtered_out_confidence_count,
                "accepted_run_count": len(runs),
                "accepted_run_pct_of_attempts": round(
                    (len(runs) / attempted_runs * 100.0) if attempted_runs > 0 else 0.0,
                    2,
                ),
                "accepted_run_pct_target": constraints.get("min_accepted_run_pct_target"),
                "accepted_run_pct_below_target": (
                    (
                        round(
                            (len(runs) / attempted_runs * 100.0)
                            if attempted_runs > 0
                            else 0.0,
                            2,
                        )
                        < float(
                            self._to_float(
                                constraints.get("min_accepted_run_pct_target")
                            )
                            or 0.0
                        )
                    )
                    if self._to_float(constraints.get("min_accepted_run_pct_target"))
                    is not None
                    else False
                ),
                "confidence_tier_hit_stats": [
                    {
                        "min_conf_exclusive": row.get("min_conf_exclusive"),
                        "max_conf_inclusive": row.get("max_conf_inclusive"),
                        "cap_pct": round(
                            (float(self._to_float(row.get("cap_frac")) or 0.0) * 100.0),
                            4,
                        ),
                        "symbol_count": int(row.get("symbol_count") or 0),
                        "hit_run_count": int(row.get("hit_run_count") or 0),
                        "hit_run_pct_of_attempts": round(
                            (
                                (int(row.get("hit_run_count") or 0) / attempted_runs) * 100.0
                                if attempted_runs > 0
                                else 0.0
                            ),
                            2,
                        ),
                        "hit_symbol_occurrences": int(
                            row.get("hit_symbol_occurrences") or 0
                        ),
                        "avg_excess_weight_pct_per_hit_run": round(
                            (
                                float(
                                    self._to_float(row.get("cumulative_excess_weight_pct"))
                                    or 0.0
                                )
                                / int(row.get("hit_run_count") or 1)
                            )
                            if int(row.get("hit_run_count") or 0) > 0
                            else 0.0,
                            4,
                        ),
                    }
                    for row in tier_hit_stats_internal
                ],
            },
            "top_scenarios": top_scenarios,
            "frontier_points": frontier_points,
        }

    def write_summary(
        self,
        summary: PositionSummary,
        output_path: Path | None = None,
    ) -> Path:
        out = output_path or (self.account_dir / f"{self.account_name}.positions.json")
        total_current_value = round(
            sum(float(p.get("current_value") or 0.0) for p in summary.positions),
            2,
        )
        total_unrealized_pl = round(
            sum(float(p.get("unrealized_pl") or 0.0) for p in summary.positions),
            2,
        )
        total_all_in_pl = round(
            sum(
                float(p.get("all_in_pl_including_dividends_and_options") or 0.0)
                for p in summary.positions
            ),
            2,
        )
        total_dividends = round(
            sum(float(p.get("dividend_sum") or 0.0) for p in summary.positions),
            2,
        )
        payload = {
            "account_name": summary.account_name,
            "records_read": summary.records_read,
            "position_count": len(summary.positions),
            "total_current_value": total_current_value,
            "total_unrealized_pl": total_unrealized_pl,
            "total_all_in_pl_including_dividends_and_options": total_all_in_pl,
            "total_dividends": total_dividends,
            "cash_reconciliation": summary.cash_reconciliation,
            "cash_forecast": summary.cash_forecast,
            "positions": summary.positions,
        }
        out.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        return out

    @staticmethod
    def _safe_sheet_name(name: str, existing: set[str]) -> str:
        base = (name or "Sheet").replace("/", "_").replace("\\", "_").replace("*", "_")
        base = base.replace("[", "(").replace("]", ")").replace(":", "-").replace("?", "")
        base = base[:31] or "Sheet"
        candidate = base
        i = 1
        while candidate in existing:
            suffix = f"_{i}"
            candidate = f"{base[: 31 - len(suffix)]}{suffix}"
            i += 1
        existing.add(candidate)
        return candidate

    def write_excel_report(
        self,
        summary: PositionSummary,
        output_path: Path | None = None,
    ) -> Path:
        out = output_path or (self.account_dir / f"{self.account_name}.positions.xlsx")
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary_headers = [
            "instrument",
            "quantity",
            "estimated_cost_basis",
            "current_price",
            "current_value",
            "dividend_sum",
            "dividend_pct_of_cost_basis",
            "all_in_pl_including_dividends_and_options",
            "total_pl_pct_of_cost_basis",
        ]
        ws_summary.append(summary_headers)
        for p in summary.positions:
            ws_summary.append(
                [
                    p.get("instrument"),
                    p.get("quantity"),
                    p.get("estimated_cost_basis"),
                    p.get("current_price"),
                    p.get("current_value"),
                    p.get("dividend_sum"),
                    p.get("dividend_pct_of_cost_basis"),
                    p.get("all_in_pl_including_dividends_and_options"),
                    p.get("total_pl_pct_of_cost_basis"),
                ]
            )

        ws_summary.append([])
        ws_summary.append(["total_positions", len(summary.positions)])
        ws_summary.append(["total_current_value", sum(float(p.get("current_value") or 0.0) for p in summary.positions)])
        ws_summary.append(["total_dividends", sum(float(p.get("dividend_sum") or 0.0) for p in summary.positions)])
        ws_summary.append(
            [
                "total_all_in_pl_including_dividends_and_options",
                sum(float(p.get("all_in_pl_including_dividends_and_options") or 0.0) for p in summary.positions),
            ]
        )
        ws_summary.append([])
        ws_summary.append(["cash_reconciliation", "value"])
        for k, v in (summary.cash_reconciliation or {}).items():
            if k == "cash_delta_by_trans_code":
                continue
            ws_summary.append([k, v])
        ws_summary.append([])
        ws_summary.append(["cash_delta_by_trans_code", "value"])
        for code, value in (summary.cash_reconciliation or {}).get(
            "cash_delta_by_trans_code", {}
        ).items():
            ws_summary.append([code, value])
        ws_summary.append([])
        ws_summary.append(["cash_forecast_as_of_date", (summary.cash_forecast or {}).get("as_of_date")])
        ws_summary.append(
            [
                "cash_forecast_projected_next_30_days",
                (summary.cash_forecast or {}).get("totals", {}).get("projected_next_30_days"),
            ]
        )
        ws_summary.append(
            [
                "cash_forecast_projected_next_60_days",
                (summary.cash_forecast or {}).get("totals", {}).get("projected_next_60_days"),
            ]
        )
        ws_summary.append(
            [
                "cash_forecast_projected_next_90_days",
                (summary.cash_forecast or {}).get("totals", {}).get("projected_next_90_days"),
            ]
        )

        # Build per-symbol full record streams once (used by ranking + symbol sheets).
        symbol_records: dict[str, list[dict[str, Any]]] = {}
        for path in sorted(self.account_dir.glob("*.json")):
            if path.name.endswith(".positions.json"):
                continue
            if path.name.endswith(".positions.xlsx"):
                continue
            rec = self._read_json(path)
            if not isinstance(rec, dict):
                continue
            symbol = str(rec.get("instrument") or "").strip().upper()
            if not symbol:
                continue
            if symbol in self.excluded_symbols:
                continue
            symbol_records.setdefault(symbol, []).append(rec)
        for symbol in symbol_records:
            symbol_records[symbol].sort(key=self._record_sort_key)

        # Precompute implied dividend % stats by symbol for ranking tab usage.
        implied_stats_by_symbol: dict[str, dict[str, float | None]] = {}
        for symbol, recs in symbol_records.items():
            running_qty = 0.0
            implied_pcts: list[float] = []
            dividend_dates: list[str] = []
            for rec in recs:
                trans_code = str(rec.get("trans_code") or "").strip().upper()
                qty = self._to_float(rec.get("quantity")) or 0.0
                amount = self._to_float(rec.get("amount")) or 0.0
                price = self._to_float(rec.get("price"))
                if trans_code in BUY_LIKE_CODES:
                    running_qty += max(0.0, qty)
                elif trans_code in SELL_LIKE_CODES:
                    running_qty = max(0.0, running_qty - max(0.0, qty))
                if trans_code not in DIVIDEND_CODES:
                    continue
                dividend_dates.append(str(rec.get("activity_date") or ""))
                div_qty_desc, div_price_desc = self._extract_div_qty_price(
                    rec.get("description")
                )
                dividend_qty = qty if qty > 0 else div_qty_desc
                _ = price if price is not None else div_price_desc  # informational only
                qty_at_div = dividend_qty if dividend_qty is not None else running_qty
                market_price_at_dividend = self._get_historical_close(
                    symbol, rec.get("activity_date")
                )
                if (
                    qty_at_div is None
                    or market_price_at_dividend is None
                    or qty_at_div <= 0
                ):
                    continue
                position_value = qty_at_div * market_price_at_dividend
                if position_value <= 0:
                    continue
                implied_pcts.append((amount / position_value) * 100)
            if implied_pcts:
                median_val = statistics.median(implied_pcts)
                if self._is_weekly_dividend_pattern(dividend_dates):
                    median_val *= 4
                implied_stats_by_symbol[symbol] = {
                    "mean_implied_dividend_pct": statistics.mean(implied_pcts),
                    "median_implied_dividend_pct": median_val,
                    "stdev_implied_dividend_pct": (
                        statistics.stdev(implied_pcts) if len(implied_pcts) > 1 else 0.0
                    ),
                }
            else:
                implied_stats_by_symbol[symbol] = {
                    "mean_implied_dividend_pct": None,
                    "median_implied_dividend_pct": None,
                    "stdev_implied_dividend_pct": None,
                }

        # Use all symbols (including zero current position) for ranking + symbol tabs.
        all_positions = summary.positions
        try:
            all_positions = self.analyze(include_zero_positions=True).positions
        except Exception:
            # Fall back to provided summary positions if all-symbol recompute fails.
            all_positions = summary.positions

        used_sheet_names = {"Summary"}

        def _add_ranked_tab(
            title: str,
            metric_key: str,
        ) -> None:
            ws = wb.create_sheet(title=self._safe_sheet_name(title, used_sheet_names))
            ws.append(
                [
                    "instrument",
                    metric_key,
                    "mean_implied_dividend_pct",
                    "median_implied_dividend_pct",
                    "stdev_implied_dividend_pct",
                    "estimated_cost_basis",
                    "metric_pct_of_cost_basis",
                    "current_value",
                    "dividend_sum",
                    "all_in_pl_including_dividends_and_options",
                ]
            )
            if title == "Dividends":
                ranked = sorted(
                    all_positions,
                    key=lambda p: float(
                        implied_stats_by_symbol.get(
                            str(p.get("instrument") or "").strip().upper(), {}
                        ).get("median_implied_dividend_pct")
                        or -1e9
                    ),
                    reverse=True,
                )
            else:
                ranked = sorted(
                    all_positions,
                    key=lambda p: float(p.get(metric_key) or 0.0),
                    reverse=True,
                )
            for p in ranked:
                cost_basis = float(p.get("estimated_cost_basis") or 0.0)
                metric_value = float(p.get(metric_key) or 0.0)
                metric_pct = (metric_value / cost_basis * 100) if cost_basis > 0 else None
                sym = str(p.get("instrument") or "").strip().upper()
                implied_stats = implied_stats_by_symbol.get(sym, {})
                ws.append(
                    [
                        p.get("instrument"),
                        p.get(metric_key),
                        (
                            round(implied_stats.get("mean_implied_dividend_pct"), 6)
                            if implied_stats.get("mean_implied_dividend_pct") is not None
                            else None
                        ),
                        (
                            round(implied_stats.get("median_implied_dividend_pct"), 6)
                            if implied_stats.get("median_implied_dividend_pct") is not None
                            else None
                        ),
                        (
                            round(implied_stats.get("stdev_implied_dividend_pct"), 6)
                            if implied_stats.get("stdev_implied_dividend_pct") is not None
                            else None
                        ),
                        p.get("estimated_cost_basis"),
                        round(metric_pct, 4) if metric_pct is not None else None,
                        p.get("current_value"),
                        p.get("dividend_sum"),
                        p.get("all_in_pl_including_dividends_and_options"),
                    ]
                )

        _add_ranked_tab("Price Appreciation", "unrealized_pl")
        _add_ranked_tab("Dividends", "dividend_sum")
        _add_ranked_tab(
            "Total Return",
            "all_in_pl_including_dividends_and_options",
        )
        position_by_symbol = {
            str(p.get("instrument") or "").strip().upper(): p for p in all_positions
        }
        symbol_universe = sorted(
            {s for s in symbol_records.keys() if s}
            | {s for s in position_by_symbol.keys() if s}
        )

        for symbol in symbol_universe:
            p = position_by_symbol.get(symbol, {})
            ws = wb.create_sheet(title=self._safe_sheet_name(symbol, used_sheet_names))

            ws.append(["symbol", symbol])
            ws.append(["quantity", p.get("quantity")])
            ws.append(["estimated_cost_basis", p.get("estimated_cost_basis")])
            ws.append(["current_price", p.get("current_price")])
            ws.append(["current_value", p.get("current_value")])
            ws.append(["dividend_sum", p.get("dividend_sum")])
            ws.append(["dividend_pct_of_cost_basis", p.get("dividend_pct_of_cost_basis")])
            ws.append(
                [
                    "all_in_pl_including_dividends_and_options",
                    p.get("all_in_pl_including_dividends_and_options"),
                ]
            )
            ws.append(["total_pl_pct_of_cost_basis", p.get("total_pl_pct_of_cost_basis")])
            ws.append([])
            ws.append(
                [
                    "activity_date",
                    "process_date",
                    "settle_date",
                    "trans_code",
                    "quantity",
                    "price",
                    "amount",
                    "description",
                ]
            )
            for txn in p.get("buy_sell_transactions") or []:
                ws.append(
                    [
                        txn.get("activity_date"),
                        txn.get("process_date"),
                        txn.get("settle_date"),
                        txn.get("trans_code"),
                        txn.get("quantity"),
                        txn.get("price"),
                        txn.get("amount"),
                        txn.get("description"),
                    ]
                )

            # Dividend events with running position context.
            ws.append([])
            ws.append(["Dividend Events"])
            ws.append(
                [
                    "activity_date",
                    "trans_code",
                    "dividend_amount",
                    "dividend_qty_shares",
                    "dividend_per_share",
                    "market_price_at_dividend",
                    "position_qty_at_dividend",
                    "position_value_at_dividend",
                    "implied_dividend_pct",
                    "running_dividend_total",
                ]
            )

            running_qty = 0.0
            running_dividend_total = 0.0
            had_dividend_row = False
            implied_pcts: list[float] = []
            dividend_dates: list[str] = []
            for rec in symbol_records.get(symbol, []):
                trans_code = str(rec.get("trans_code") or "").strip().upper()
                qty = self._to_float(rec.get("quantity")) or 0.0
                amount = self._to_float(rec.get("amount")) or 0.0
                price = self._to_float(rec.get("price"))

                if trans_code in BUY_LIKE_CODES:
                    running_qty += max(0.0, qty)
                elif trans_code in SELL_LIKE_CODES:
                    running_qty = max(0.0, running_qty - max(0.0, qty))

                if trans_code in DIVIDEND_CODES:
                    had_dividend_row = True
                    dividend_dates.append(str(rec.get("activity_date") or ""))
                    running_dividend_total += amount
                    div_qty_desc, div_price_desc = self._extract_div_qty_price(
                        rec.get("description")
                    )
                    dividend_qty = qty if qty > 0 else div_qty_desc
                    dividend_per_share = price if price is not None else div_price_desc
                    # Prefer explicit dividend-share qty when present.
                    qty_at_div = dividend_qty if dividend_qty is not None else running_qty
                    market_price_at_dividend = self._get_historical_close(
                        symbol, rec.get("activity_date")
                    )
                    position_value_at_div = (
                        qty_at_div * market_price_at_dividend
                        if (
                            qty_at_div is not None
                            and market_price_at_dividend is not None
                            and qty_at_div > 0
                        )
                        else None
                    )
                    implied_pct = (
                        (amount / position_value_at_div) * 100
                        if (position_value_at_div is not None and position_value_at_div > 0)
                        else None
                    )
                    if implied_pct is not None:
                        implied_pcts.append(implied_pct)
                    ws.append(
                        [
                            rec.get("activity_date"),
                            rec.get("trans_code"),
                            amount,
                            dividend_qty,
                            dividend_per_share,
                            market_price_at_dividend,
                            round(qty_at_div, 6) if qty_at_div is not None else None,
                            round(position_value_at_div, 2)
                            if position_value_at_div is not None
                            else None,
                            round(implied_pct, 4) if implied_pct is not None else None,
                            round(running_dividend_total, 2),
                        ]
                    )

            if not had_dividend_row:
                ws.append(["(no dividend events)"])
            ws.append([])
            ws.append(["Implied Dividend % Stats"])
            if implied_pcts:
                mean_val = statistics.mean(implied_pcts)
                median_val = statistics.median(implied_pcts)
                if self._is_weekly_dividend_pattern(dividend_dates):
                    median_val *= 4
                stdev_val = (
                    statistics.stdev(implied_pcts) if len(implied_pcts) > 1 else 0.0
                )
                ws.append(["mean_implied_dividend_pct", round(mean_val, 6)])
                ws.append(["median_implied_dividend_pct", round(median_val, 6)])
                ws.append(["stdev_implied_dividend_pct", round(stdev_val, 6)])
            else:
                ws.append(["mean_implied_dividend_pct", None])
                ws.append(["median_implied_dividend_pct", None])
                ws.append(["stdev_implied_dividend_pct", None])

        # Monthly summary tabs based on imported account records.
        month_symbol: dict[str, dict[str, dict[str, float]]] = {}
        for path in sorted(self.account_dir.glob("*.json")):
            if path.name.endswith(".positions.json"):
                continue
            if path.name.endswith(".positions.xlsx"):
                continue
            rec = self._read_json(path)
            if not isinstance(rec, dict):
                continue
            date_s = str(rec.get("activity_date") or "").strip()
            if len(date_s) < 7:
                continue
            month_key = date_s[:7]
            symbol = str(rec.get("instrument") or "").strip().upper() or "(NO_SYMBOL)"
            if symbol != "(NO_SYMBOL)" and symbol in self.excluded_symbols:
                continue
            trans_code = str(rec.get("trans_code") or "").strip().upper()
            qty = self._to_float(rec.get("quantity")) or 0.0
            amount = self._to_float(rec.get("amount")) or 0.0

            month_bucket = month_symbol.setdefault(month_key, {})
            symbol_bucket = month_bucket.setdefault(
                symbol,
                {
                    "buy_like_qty": 0.0,
                    "sell_like_qty": 0.0,
                    "net_position_change_qty": 0.0,
                    "dividends_amount": 0.0,
                    "position_change_cash_net": 0.0,
                },
            )

            if trans_code in BUY_LIKE_CODES:
                symbol_bucket["buy_like_qty"] += qty
                symbol_bucket["net_position_change_qty"] += qty
            elif trans_code in SELL_LIKE_CODES:
                symbol_bucket["sell_like_qty"] += qty
                symbol_bucket["net_position_change_qty"] -= qty

            if trans_code in DIVIDEND_CODES:
                symbol_bucket["dividends_amount"] += amount

            if trans_code in MONTHLY_POSITION_CHANGE_CODES:
                symbol_bucket["position_change_cash_net"] += amount

        for month_key in sorted(month_symbol.keys()):
            ws_month = wb.create_sheet(
                title=self._safe_sheet_name(month_key, used_sheet_names)
            )
            ws_month.append(
                [
                    "instrument",
                    "buy_like_qty",
                    "sell_like_qty",
                    "net_position_change_qty",
                    "dividends_amount",
                    "position_change_cash_net",
                ]
            )

            total_dividends = 0.0
            total_net_position_change = 0.0
            total_position_change_cash_net = 0.0

            for symbol in sorted(month_symbol[month_key].keys()):
                row = month_symbol[month_key][symbol]
                total_dividends += row["dividends_amount"]
                total_net_position_change += row["net_position_change_qty"]
                total_position_change_cash_net += row["position_change_cash_net"]
                ws_month.append(
                    [
                        symbol,
                        round(row["buy_like_qty"], 6),
                        round(row["sell_like_qty"], 6),
                        round(row["net_position_change_qty"], 6),
                        round(row["dividends_amount"], 2),
                        round(row["position_change_cash_net"], 2),
                    ]
                )

            ws_month.append([])
            ws_month.append(["month", month_key])
            ws_month.append(["total_dividends", round(total_dividends, 2)])
            ws_month.append(
                [
                    "total_net_position_change_qty",
                    round(total_net_position_change, 6),
                ]
            )
            ws_month.append(
                [
                    "total_position_change_cash_net",
                    round(total_position_change_cash_net, 2),
                ]
            )

        # Cash forecast tab.
        ws_cf = wb.create_sheet(title=self._safe_sheet_name("Cash Forecast", used_sheet_names))
        cash_forecast = summary.cash_forecast or {}
        ws_cf.append(["as_of_date", cash_forecast.get("as_of_date")])
        ws_cf.append(
            [
                "projected_next_30_days",
                (cash_forecast.get("totals") or {}).get("projected_next_30_days"),
            ]
        )
        ws_cf.append(
            [
                "projected_next_60_days",
                (cash_forecast.get("totals") or {}).get("projected_next_60_days"),
            ]
        )
        ws_cf.append(
            [
                "projected_next_90_days",
                (cash_forecast.get("totals") or {}).get("projected_next_90_days"),
            ]
        )
        ws_cf.append([])
        cash_forecast_headers = [
            "instrument",
            "cadence_label",
            "cadence_days",
            "events_observed",
            "last_dividend_date",
            "median_dividend_amount",
            "shares_used_for_projection",
            "projected_next_30_days",
            "projected_growth_dollars",
            "projected_next_60_days",
            "projected_next_90_days",
            "projected_payout_dates",
            "forecast_note",
            "yield_source",
            "dividend_yield_used",
            "per_share_distribution_rate_used",
            "effective_30_day_distribution_rate",
            "current_value",
            "estimated_cost_basis",
            "total_pl_pct_of_cost_basis",
            "income_optimized_target_shares",
            "income_optimized_target_30_day_distribution",
            "income_optimized_target_projected_growth_dollars",
            "growth_optimized_target_shares",
            "growth_optimized_target_30_day_distribution",
            "growth_optimized_target_projected_growth_dollars",
            "balanced_optimized_target_shares",
            "balanced_optimized_target_30_day_distribution",
            "balanced_optimized_target_projected_growth_dollars",
        ]
        ws_cf.append(cash_forecast_headers)
        for row in cash_forecast.get("symbols") or []:
            ws_cf.append(
                [
                    row.get("instrument"),
                    row.get("cadence_label"),
                    row.get("cadence_days"),
                    row.get("events_observed"),
                    row.get("last_dividend_date"),
                    row.get("median_dividend_amount"),
                    row.get("shares_used_for_projection"),
                    row.get("projected_next_30_days"),
                    row.get("projected_growth_dollars"),
                    row.get("projected_next_60_days"),
                    row.get("projected_next_90_days"),
                    ", ".join(row.get("projected_payout_dates") or []),
                    row.get("forecast_note"),
                    row.get("yield_source"),
                    row.get("dividend_yield_used"),
                    row.get("per_share_distribution_rate_used"),
                    row.get("effective_30_day_distribution_rate"),
                    row.get("current_value"),
                    row.get("estimated_cost_basis"),
                    row.get("total_pl_pct_of_cost_basis"),
                    row.get("income_optimized_target_shares"),
                    row.get("income_optimized_target_30_day_distribution"),
                    row.get("income_optimized_target_projected_growth_dollars"),
                    row.get("growth_optimized_target_shares"),
                    row.get("growth_optimized_target_30_day_distribution"),
                    row.get("growth_optimized_target_projected_growth_dollars"),
                    row.get("balanced_optimized_target_shares"),
                    row.get("balanced_optimized_target_30_day_distribution"),
                    row.get("balanced_optimized_target_projected_growth_dollars"),
                ]
            )

        # Totals row for share/distribution/growth dollar columns.
        total_keys = {
            "median_dividend_amount",
            "projected_next_30_days",
            "projected_growth_dollars",
            "projected_next_60_days",
            "projected_next_90_days",
            "shares_used_for_projection",
            "current_value",
            "estimated_cost_basis",
            "income_optimized_target_shares",
            "income_optimized_target_30_day_distribution",
            "income_optimized_target_projected_growth_dollars",
            "growth_optimized_target_shares",
            "growth_optimized_target_30_day_distribution",
            "growth_optimized_target_projected_growth_dollars",
            "balanced_optimized_target_shares",
            "balanced_optimized_target_30_day_distribution",
            "balanced_optimized_target_projected_growth_dollars",
        }
        total_map: dict[str, float] = {k: 0.0 for k in total_keys}
        for row in cash_forecast.get("symbols") or []:
            for key in total_keys:
                val = self._to_float(row.get(key))
                if val is not None:
                    total_map[key] += val
        ws_cf.append(
            [
                "TOTAL",
                None,
                None,
                None,
                None,
                round(total_map["median_dividend_amount"], 2),
                round(total_map["shares_used_for_projection"], 6),
                round(total_map["projected_next_30_days"], 2),
                round(total_map["projected_growth_dollars"], 2),
                round(total_map["projected_next_60_days"], 2),
                round(total_map["projected_next_90_days"], 2),
                None,
                None,
                None,
                None,
                None,
                None,
                round(total_map["current_value"], 2),
                round(total_map["estimated_cost_basis"], 2),
                None,
                round(total_map["income_optimized_target_shares"], 6),
                round(total_map["income_optimized_target_30_day_distribution"], 2),
                round(total_map["income_optimized_target_projected_growth_dollars"], 2),
                round(total_map["growth_optimized_target_shares"], 6),
                round(total_map["growth_optimized_target_30_day_distribution"], 2),
                round(total_map["growth_optimized_target_projected_growth_dollars"], 2),
                round(total_map["balanced_optimized_target_shares"], 6),
                round(total_map["balanced_optimized_target_30_day_distribution"], 2),
                round(total_map["balanced_optimized_target_projected_growth_dollars"], 2),
            ]
        )

        # Explain optimization pair logic after the symbol list.
        ws_cf.append([])
        ws_cf.append(["Optimization Pair Explanations"])
        ws_cf.append(
            [
                "income_optimized_target_shares",
                (
                    "Score = effective_30_day_distribution_rate (non-negative). "
                    "Target value weight = symbol score / sum(scores). "
                    "Target shares = target value / current price."
                ),
            ]
        )
        ws_cf.append(
            [
                "income_optimized_target_30_day_distribution",
                (
                    "Target 30-day distribution = income_optimized_target_shares * "
                    "effective_30_day_distribution_rate."
                ),
            ]
        )
        ws_cf.append(
            [
                "growth_optimized_target_shares",
                (
                    "Score = total_pl_pct_of_cost_basis (non-negative). "
                    "Target value weight = symbol score / sum(scores). "
                    "Target shares = target value / current price."
                ),
            ]
        )
        ws_cf.append(
            [
                "growth_optimized_target_30_day_distribution",
                (
                    "Target 30-day distribution = growth_optimized_target_shares * "
                    "effective_30_day_distribution_rate."
                ),
            ]
        )
        ws_cf.append(
            [
                "balanced_optimized_target_shares",
                (
                    "Score = average(income score, growth score). "
                    "Income score = effective_30_day_distribution_rate (non-negative). "
                    "Growth score = total_pl_pct_of_cost_basis (non-negative). "
                    "Target value weight = symbol score / sum(scores). "
                    "Target shares = target value / current price."
                ),
            ]
        )
        ws_cf.append(
            [
                "balanced_optimized_target_30_day_distribution",
                (
                    "Target 30-day distribution = balanced_optimized_target_shares * "
                    "effective_30_day_distribution_rate."
                ),
            ]
        )

        # Write a standalone Cash Forecast workbook copy.
        cash_forecast_out = out.with_name(f"{self.account_name}.cash_forecast.xlsx")
        cf_wb = Workbook()
        cf_ws = cf_wb.active
        cf_ws.title = "Cash Forecast"
        for row in ws_cf.iter_rows(values_only=True):
            cf_ws.append(list(row))

        # Add Monte Carlo analysis tab for growth/income optimization areas.
        mc = self._build_cash_monte_carlo_analysis(cash_forecast.get("symbols") or [])
        mc_ws = cf_wb.create_sheet(title="Monte Carlo")
        mc_ws.append(["simulations", mc.get("simulations")])
        mc_ws.append(["attempted_simulations", mc.get("attempted_simulations")])
        mc_ws.append(["portfolio_current_value", mc.get("total_value")])
        mc_ws.append(["symbols_in_universe", mc.get("symbol_count")])
        mc_constraints = mc.get("constraints_applied") or {}
        mc_ws.append(["simulation_count", mc_constraints.get("simulation_count")])
        mc_ws.append(["sampling_alpha_scale", mc_constraints.get("sampling_alpha_scale")])
        mc_ws.append(
            [
                "include_symbols_even_if_not_held",
                json.dumps(
                    mc_constraints.get("include_symbols_even_if_not_held") or [],
                    ensure_ascii=False,
                ),
            ]
        )
        mc_ws.append(["max_symbol_weight_pct", mc_constraints.get("max_symbol_weight_pct")])
        mc_ws.append(["max_top_5_weight_pct", mc_constraints.get("max_top_5_weight_pct")])
        mc_ws.append(["max_top_10_weight_pct", mc_constraints.get("max_top_10_weight_pct")])
        mc_ws.append(
            [
                "confidence_tier_symbol_caps",
                json.dumps(
                    mc_constraints.get("confidence_tier_symbol_caps") or [],
                    ensure_ascii=False,
                ),
            ]
        )
        mc_ws.append(
            [
                "high_confidence_symbol_cap",
                json.dumps(
                    mc_constraints.get("high_confidence_symbol_cap") or {},
                    ensure_ascii=False,
                ),
            ]
        )
        mc_ws.append(
            [
                "max_income_contribution_pct",
                mc_constraints.get("max_income_contribution_pct"),
            ]
        )
        mc_ws.append(
            [
                "max_growth_contribution_pct",
                mc_constraints.get("max_growth_contribution_pct"),
            ]
        )
        mc_ws.append(
            [
                "allocation_confidence_growth_weight",
                mc_constraints.get("allocation_confidence_growth_weight"),
            ]
        )
        mc_ws.append(
            [
                "allocation_confidence_dividend_weight",
                mc_constraints.get("allocation_confidence_dividend_weight"),
            ]
        )
        mc_ws.append(
            [
                "allocation_confidence_multiplier_min",
                mc_constraints.get("allocation_confidence_multiplier_min"),
            ]
        )
        mc_ws.append(
            [
                "allocation_confidence_multiplier_max",
                mc_constraints.get("allocation_confidence_multiplier_max"),
            ]
        )
        mc_ws.append(
            [
                "min_accepted_run_pct_target",
                mc_constraints.get("min_accepted_run_pct_target"),
            ]
        )
        mc_ws.append(
            [
                "min_price_history_months",
                mc_constraints.get("min_price_history_months"),
            ]
        )
        mc_ws.append(
            [
                "min_confidence_score",
                mc_constraints.get("min_confidence_score"),
            ]
        )
        mc_ws.append(
            [
                "low_confidence_score_threshold",
                mc_constraints.get("low_confidence_score_threshold"),
            ]
        )
        mc_ws.append(
            [
                "max_low_confidence_total_weight_pct",
                mc_constraints.get("max_low_confidence_total_weight_pct"),
            ]
        )
        mc_ws.append(
            [
                "max_annualized_growth_pct_for_projection",
                mc_constraints.get("max_annualized_growth_pct_for_projection"),
            ]
        )
        mc_ws.append(
            [
                "top_concentration_shocks_pct",
                json.dumps(
                    mc_constraints.get("top_concentration_shocks_pct") or [],
                    ensure_ascii=False,
                ),
            ]
        )
        for cap in (mc_constraints.get("theme_caps") or []):
            if not isinstance(cap, dict):
                continue
            mc_ws.append(
                [
                    f"theme_cap:{cap.get('name')}",
                    cap.get("max_weight_pct"),
                    ",".join(str(s or "").strip().upper() for s in (cap.get("symbols") or [])),
                ]
            )
        mc_hits = mc.get("constraint_hit_stats") or {}
        mc_ws.append([])
        mc_ws.append(["constraint_hit_stats", "count", "pct_of_attempts"])
        mc_ws.append(
            [
                "symbol_cap_hit",
                mc_hits.get("symbol_cap_hit_count"),
                mc_hits.get("symbol_cap_hit_pct_of_attempts"),
            ]
        )
        mc_ws.append(
            [
                "theme_cap_hit",
                mc_hits.get("theme_cap_hit_count"),
                mc_hits.get("theme_cap_hit_pct_of_attempts"),
            ]
        )
        mc_ws.append(
            [
                "rejected_growth_contribution",
                mc_hits.get("rejected_growth_contribution_count"),
                mc_hits.get("rejected_growth_contribution_pct_of_attempts"),
            ]
        )
        mc_ws.append(
            [
                "rejected_income_contribution",
                mc_hits.get("rejected_income_contribution_count"),
                mc_hits.get("rejected_income_contribution_pct_of_attempts"),
            ]
        )
        mc_ws.append(
            [
                "rejected_low_confidence_weight",
                mc_hits.get("rejected_low_confidence_weight_count"),
                mc_hits.get("rejected_low_confidence_weight_pct_of_attempts"),
            ]
        )
        mc_ws.append(
            [
                "rejected_top_5_weight",
                mc_hits.get("rejected_top_5_weight_count"),
                mc_hits.get("rejected_top_5_weight_pct_of_attempts"),
            ]
        )
        mc_ws.append(
            [
                "rejected_top_10_weight",
                mc_hits.get("rejected_top_10_weight_count"),
                mc_hits.get("rejected_top_10_weight_pct_of_attempts"),
            ]
        )
        mc_ws.append(
            [
                "filtered_out_min_price_history",
                mc_hits.get("filtered_out_min_price_history_count"),
                "",
            ]
        )
        mc_ws.append(
            [
                "filtered_out_min_confidence",
                mc_hits.get("filtered_out_min_confidence_count"),
                "",
            ]
        )
        mc_ws.append(
            [
                "accepted_runs",
                mc_hits.get("accepted_run_count"),
                mc_hits.get("accepted_run_pct_of_attempts"),
            ]
        )
        mc_ws.append(
            [
                "accepted_run_pct_target",
                "",
                mc_hits.get("accepted_run_pct_target"),
            ]
        )
        mc_ws.append(
            [
                "accepted_run_pct_below_target",
                "",
                mc_hits.get("accepted_run_pct_below_target"),
            ]
        )
        tier_stats = mc_hits.get("confidence_tier_hit_stats") or []
        if isinstance(tier_stats, list) and tier_stats:
            mc_ws.append([])
            mc_ws.append(
                [
                    "confidence_tier_hit_stats",
                    "symbol_count",
                    "hit_run_count",
                    "hit_run_pct_of_attempts",
                    "hit_symbol_occurrences",
                    "avg_excess_weight_pct_per_hit_run",
                ]
            )
            for row in tier_stats:
                if not isinstance(row, dict):
                    continue
                min_conf = self._to_float(row.get("min_conf_exclusive"))
                max_conf = self._to_float(row.get("max_conf_inclusive"))
                cap_pct = self._to_float(row.get("cap_pct"))
                if min_conf is None:
                    band = f"<= {max_conf} @ cap {cap_pct}%"
                else:
                    band = f"> {min_conf} and <= {max_conf} @ cap {cap_pct}%"
                mc_ws.append(
                    [
                        band,
                        row.get("symbol_count"),
                        row.get("hit_run_count"),
                        row.get("hit_run_pct_of_attempts"),
                        row.get("hit_symbol_occurrences"),
                        row.get("avg_excess_weight_pct_per_hit_run"),
                    ]
                )
        mc_ws.append([])
        mc_ws.append(
            [
                "scenario",
                "expected_30_day_distribution",
                "expected_growth_dollars",
                "allocation_top_weights",
                "top_5_weight_pct",
                "top_10_weight_pct",
                "top_concentration_stress",
            ]
        )
        for row in mc.get("top_scenarios") or []:
            mc_ws.append(
                [
                    row.get("scenario"),
                    row.get("expected_30_day_distribution"),
                    row.get("expected_growth_dollars"),
                    row.get("allocation_top_weights"),
                    row.get("top_5_weight_pct"),
                    row.get("top_10_weight_pct"),
                    json.dumps(row.get("top_concentration_stress") or {}, ensure_ascii=False),
                ]
            )
        mc_ws.append([])
        mc_ws.append(
            [
                "income_decile",
                "expected_30_day_distribution",
                "expected_growth_dollars",
                "allocation_top_weights",
            ]
        )
        for row in mc.get("frontier_points") or []:
            mc_ws.append(
                [
                    row.get("income_decile"),
                    row.get("expected_30_day_distribution"),
                    row.get("expected_growth_dollars"),
                    row.get("allocation_top_weights"),
                ]
            )

        # Add one tab per scenario with full allocation breakdown.
        cf_sheet_names = set(cf_wb.sheetnames)
        for row in mc.get("top_scenarios") or []:
            title = self._safe_sheet_name(
                f"Scenario {row.get('scenario')}", cf_sheet_names
            )
            ws_s = cf_wb.create_sheet(title=title)
            ws_s.append(["scenario", row.get("scenario")])
            ws_s.append(
                [
                    "expected_30_day_distribution",
                    row.get("expected_30_day_distribution"),
                ]
            )
            ws_s.append(["expected_growth_dollars", row.get("expected_growth_dollars")])
            ws_s.append(
                [
                    "scenario_constraint_unmet_fallback",
                    row.get("scenario_constraint_unmet_fallback"),
                ]
            )
            ws_s.append(
                [
                    "scenario_constraints_applied",
                    json.dumps(row.get("scenario_constraints_applied") or {}, ensure_ascii=False),
                ]
            )
            ws_s.append(["top_5_weight_pct", row.get("top_5_weight_pct")])
            ws_s.append(["top_10_weight_pct", row.get("top_10_weight_pct")])
            ws_s.append(
                [
                    "top_concentration_stress",
                    json.dumps(row.get("top_concentration_stress") or {}, ensure_ascii=False),
                ]
            )
            ws_s.append([])
            ws_s.append(
                [
                    "instrument",
                    "weight_pct",
                    "target_value",
                    "target_shares",
                    "expected_30_day_distribution",
                    "expected_growth_dollars",
                    "price_history_months",
                    "derived_price_growth_pct",
                    "price_history_confidence_score",
                    "dividend_history_rate_30_day_per_share",
                    "dividend_history_confidence_score",
                    "dividend_history_months",
                    "dividend_history_source",
                    "allocation_confidence_score",
                    "allocation_confidence_multiplier",
                ]
            )
            for alloc in row.get("allocation_rows") or []:
                ws_s.append(
                    [
                        alloc.get("instrument"),
                        alloc.get("weight_pct"),
                        alloc.get("target_value"),
                        alloc.get("target_shares"),
                        alloc.get("expected_30_day_distribution"),
                        alloc.get("expected_growth_dollars"),
                        alloc.get("price_history_months"),
                        alloc.get("derived_price_growth_pct"),
                        alloc.get("price_history_confidence_score"),
                        alloc.get("dividend_history_rate_30_day_per_share"),
                        alloc.get("dividend_history_confidence_score"),
                        alloc.get("dividend_history_months"),
                        alloc.get("dividend_history_source"),
                        alloc.get("allocation_confidence_score"),
                        alloc.get("allocation_confidence_multiplier"),
                    ]
                )

        # Add one tab per income decile frontier point with full allocation breakdown.
        for row in mc.get("frontier_points") or []:
            decile = row.get("income_decile")
            title = self._safe_sheet_name(f"Decile {decile}", cf_sheet_names)
            ws_d = cf_wb.create_sheet(title=title)
            ws_d.append(["income_decile", decile])
            ws_d.append(
                [
                    "expected_30_day_distribution",
                    row.get("expected_30_day_distribution"),
                ]
            )
            ws_d.append(["expected_growth_dollars", row.get("expected_growth_dollars")])
            ws_d.append([])
            ws_d.append(
                [
                    "instrument",
                    "weight_pct",
                    "target_value",
                    "target_shares",
                    "expected_30_day_distribution",
                    "expected_growth_dollars",
                    "price_history_months",
                    "derived_price_growth_pct",
                    "price_history_confidence_score",
                    "dividend_history_rate_30_day_per_share",
                    "dividend_history_confidence_score",
                    "dividend_history_months",
                    "dividend_history_source",
                    "allocation_confidence_score",
                    "allocation_confidence_multiplier",
                ]
            )
            for alloc in row.get("allocation_rows") or []:
                ws_d.append(
                    [
                        alloc.get("instrument"),
                        alloc.get("weight_pct"),
                        alloc.get("target_value"),
                        alloc.get("target_shares"),
                        alloc.get("expected_30_day_distribution"),
                        alloc.get("expected_growth_dollars"),
                        alloc.get("price_history_months"),
                        alloc.get("derived_price_growth_pct"),
                        alloc.get("price_history_confidence_score"),
                        alloc.get("dividend_history_rate_30_day_per_share"),
                        alloc.get("dividend_history_confidence_score"),
                        alloc.get("dividend_history_months"),
                        alloc.get("dividend_history_source"),
                        alloc.get("allocation_confidence_score"),
                        alloc.get("allocation_confidence_multiplier"),
                    ]
                )
        cf_wb.save(cash_forecast_out)

        wb.save(out)
        return out


__all__ = ["PositionSummary", "RobinhoodPositionAnalyzer"]
