"""Workbook discovery and inspection helpers for portfolio intent agent."""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook


ProgressFn = Callable[[str], None]


def is_valid_xlsx(path: Path) -> tuple[bool, str | None]:
    if not path.is_file():
        return False, "file does not exist"
    if not zipfile.is_zipfile(path):
        return False, "not a valid xlsx zip container"
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        _ = wb.sheetnames
        return True, None
    except Exception as exc:
        return False, str(exc)


def find_latest_valid_cash_forecast_xlsx(repo_root: Path) -> tuple[Path | None, list[str]]:
    base = repo_root / "application_files" / "data" / "portfolio" / "robinhood"
    if not base.is_dir():
        return None, []
    candidates = [
        path
        for path in base.rglob("*.cash_forecast.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    diagnostics: list[str] = []
    for candidate in candidates:
        ok, reason = is_valid_xlsx(candidate)
        if ok:
            return candidate, diagnostics
        diagnostics.append(f"{candidate}: {reason}")
    return None, diagnostics


def resolve_workbook_path(raw_path: Any, default_workbook: Path | None) -> Path:
    path = Path(raw_path).expanduser().resolve() if raw_path else default_workbook
    if path is None:
        raise RuntimeError("No workbook path provided and no default found.")
    ok, reason = is_valid_xlsx(path)
    if not ok:
        raise RuntimeError(f"Workbook is not a valid .xlsx file: {path} ({reason})")
    return path


def sheet_rows(
    path: Path,
    sheet_name: str,
    *,
    max_rows: int = 120,
    progress: ProgressFn | None = None,
) -> list[list[Any]]:
    if progress:
        progress(f"Reading sheet '{sheet_name}' from {path.name} (max_rows={max_rows})")
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    rows: list[list[Any]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx > max_rows:
            break
        values = list(row)
        if any(v is not None and str(v).strip() != "" for v in values):
            rows.append(values)
    return rows


def inspect_monte_carlo(path: Path, *, progress: ProgressFn | None = None) -> dict[str, Any]:
    if progress:
        progress(f"Inspecting Monte Carlo workbook: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = wb.sheetnames
    monte_sheet = "Monte Carlo" if "Monte Carlo" in sheets else None
    scenario_sheets = [name for name in sheets if name.lower().startswith("scenario ")]
    decile_sheets = [name for name in sheets if name.lower().startswith("decile ")]
    out: dict[str, Any] = {
        "workbook_path": str(path),
        "sheetnames": sheets,
        "scenario_sheet_count": len(scenario_sheets),
        "decile_sheet_count": len(decile_sheets),
        "scenario_sheets": scenario_sheets,
        "decile_sheets": decile_sheets,
    }
    if monte_sheet:
        out["monte_carlo_preview"] = sheet_rows(path, monte_sheet, max_rows=140, progress=progress)
    if scenario_sheets:
        out["top_scenario_preview"] = sheet_rows(
            path, scenario_sheets[0], max_rows=80, progress=progress
        )
    if decile_sheets:
        out["top_decile_preview"] = sheet_rows(path, decile_sheets[0], max_rows=80, progress=progress)
    return out

